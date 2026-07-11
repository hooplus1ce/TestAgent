"""Tests for the JSON-to-Excel delivery artifact."""
import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook


def _load_exporter():
    path = Path(__file__).parents[1] / ".claude/skills/test-case-generator-dp/scripts/generate_from_json.py"
    spec = importlib.util.spec_from_file_location("generate_from_json", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_excel_auxiliary_sheet_uses_real_asset_inventory(tmp_path):
    exporter = _load_exporter()
    payload = {
        "module_info": {
            "enterprise_prefix": "NB", "module_pinyin": "GZMX",
            "module_name": "生产管理_工资明细", "module_level1": "生产管理",
            "module_level2": "工资明细", "author": "Tester",
        },
        "test_cases": [{
            "case_id": "I001", "case_title": "编辑工资明细", "priority": "中级",
            "test_type": "功能测试", "function": "编辑", "verify_point": "金额联动",
            "preconditions": ["已登录"], "test_steps": ["编辑单价"],
            "test_data": {"备注": "FIXTURE-1"}, "expected_result": "金额更新",
            "automation_suggestion": "自动回放",
        }],
        "asset_inventory": [
            {"asset_type": "filter", "name": "备注", "metadata": {"value_mode": "free-text"}},
            {"asset_type": "action", "name": "编辑", "metadata": {"area": "页面", "kind": "primary"}},
            {"asset_type": "table_column", "name": "金额", "metadata": {"row_count": 3}},
        ],
    }
    source = tmp_path / "suite.json"
    source.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    module_info, cases, assets = exporter.load_testcases_from_json([str(source)])
    target = exporter.build_excel(
        module_info, cases, output_dir=str(tmp_path), custom_filename="cases.xlsx",
        asset_inventory=assets,
    )

    workbook = load_workbook(target, read_only=True, data_only=True)
    sheet = workbook["测试数据"]
    values = [str(cell.value or "") for row in sheet.iter_rows() for cell in row]
    workbook.close()
    assert "备注" in values and "编辑" in values and "金额" in values
    assert not any("探索过程中自动填充" in value for value in values)
    assert not any("['" in value for value in values)
