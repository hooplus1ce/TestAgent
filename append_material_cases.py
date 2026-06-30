#!/usr/bin/env python3
"""追加按物料筛选查询用例到现有Excel"""
import os
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PREFIX, AUTHOR = "NB", "Hooplus1ce"
L1, L2, PIN = "生产管理", "制令单明细表", "ZLDMB"
TODAY = date.today().isoformat()

# 加载已有Excel
src = "screenshots/测试用例_生产管理_制令单明细表_筛选查询_2026-06-30.xlsx"
wb = load_workbook(src)
ws = wb["测试用例"]
ws2 = wb["测试数据"]

tb = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
PRIO = {
    "高级": (PatternFill(start_color="FFD2D2",end_color="FFD2D2",fill_type="solid"), Font(bold=True,color="9C0006",size=10,name="微软雅黑")),
    "中级": (PatternFill(start_color="FFE8D0",end_color="FFE8D0",fill_type="solid"), Font(bold=True,color="9C6500",size=10,name="微软雅黑")),
    "低级": (PatternFill(start_color="FFF5C0",end_color="FFF5C0",fill_type="solid"), Font(bold=True,color="806000",size=10,name="微软雅黑")),
}
def prio(cell):
    s = PRIO.get(cell.value)
    if s: cell.fill, cell.font = s

PRE2 = "1. 已登录诺贝科技SCM系统\n2. 在「生产管理-制令单明细表」列表页\n3. 已切换至「按物料」视图\n4. 表格已加载物料数据"

def tc(n,t,l,v,ty,s,j,d,exp):
    return [n,t,l,v,L1,L2,ty,s,PRE2,j,d,exp,"","","",AUTHOR,TODAY,""]

TC2 = [
 tc(f"{PREFIX}_{PIN}_F030","筛选(物料)-物料编码包含3001","中级","验证物料编码文本包含筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料编码」输入「3001」\n3. 点击查询","物料编码包含3001",
    "表格所有物料编码均包含「3001」"),

 tc(f"{PREFIX}_{PIN}_F031","筛选(物料)-物料名称包含艾玛","中级","验证物料名称文本包含筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料名称」输入「艾玛」\n3. 点击查询","物料名称包含艾玛",
    "表格所有物料名称均包含「艾玛」"),

 tc(f"{PREFIX}_{PIN}_F032","筛选(物料)-仓库等于半成品仓","中级","验证仓库文本包含筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「仓库」输入「半成品仓」\n3. 点击查询","仓库包含半成品仓",
    "表格所有仓库均包含「半成品仓」"),

 tc(f"{PREFIX}_{PIN}_F033","筛选(物料)-应发数量等于1","中级","验证应发数量等于筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「应发数量」输入「1」\n3. 点击查询","应发数量=1",
    "表格所有记录的应发数量均为1"),

 tc(f"{PREFIX}_{PIN}_F034","筛选(物料)-物料名称包含中型胶垫圈","中级","验证物料名称筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料名称」输入「中型胶垫圈」\n3. 点击查询","物料名称包含中型胶垫圈",
    "表格所有物料名称均包含「中型胶垫圈」"),

 tc(f"{PREFIX}_{PIN}_F035","筛选(物料)-物料单位包含个","低级","验证物料单位文本包含筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料单位」输入「个」\n3. 点击查询","物料单位包含个",
    "表格所有物料单位均包含「个」"),

 tc(f"{PREFIX}_{PIN}_F036","筛选(物料)-制造排产等于已排产","中级","验证物料视图下状态筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「制造排产状态」选「已排产」\n3. 点击查询","制造排产状态=已排产",
    "表格记录的制造排产状态均为「已排产」"),

 tc(f"{PREFIX}_{PIN}_F037","筛选(物料)-组合:物料名称+仓库","高级","验证物料+仓库组合筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料名称」输入「艾玛」\n3. 「仓库」输入「半成品仓」\n4. 点击查询",
    "物料名称含艾玛+仓库含半成品仓",
    "表格记录同时满足：物料名称含「艾玛」、仓库含「半成品仓」"),

 tc(f"{PREFIX}_{PIN}_F038","筛选(物料)-排产开始日期范围","中级","验证排产开始日期范围筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「排产开始日期」选择范围\n3. 点击查询","排产开始日期=日期范围",
    "表格记录的排产开始日期在所选范围内"),

 tc(f"{PREFIX}_{PIN}_F039","筛选(物料)-应发数量不等于1","中级","验证应发数量不等于筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「应发数量」操作符选「不等于」\n3. 输入「1」\n4. 点击查询","应发数量≠1",
    "表格记录的应发数量均不为1"),

 tc(f"{PREFIX}_{PIN}_F040","筛选(物料)-物料编码为空","低级","验证物料编码为空筛选","功能","筛选查询",
    "1. 展开筛选区\n2. 「物料编码」操作符选「为空」\n3. 点击查询","物料编码为空",
    "表格显示物料编码为空的记录"),
]

next_row = ws.max_row + 1
for i, rd in enumerate(TC2):
    for ci, v in enumerate(rd, 1):
        cell = ws.cell(row=next_row+i, column=ci, value=v)
        cell.border = tb
        cell.font = Font(size=10, name="微软雅黑")
        if ci in (2,4,9,10,11,12):
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    prio(ws.cell(row=next_row+i, column=3))

# 更新自动筛选范围
ws.auto_filter.ref = f"A1:R{ws.max_row-1}"

# 追加到测试数据Sheet
sf = PatternFill(start_color="D9E2F3",end_color="D9E2F3",fill_type="solid")
sfn = Font(bold=True,size=11,name="微软雅黑")
h2f = Font(bold=True,size=10,name="微软雅黑")

def sec(ws,sr,title,hdrs,rows_data):
    ws.cell(row=sr,column=1,value=title).font=sfn
    for c in range(1,len(hdrs)+1):
        ws.cell(row=sr,column=c).fill=sf; ws.cell(row=sr,column=c).font=sfn
    hr=sr+1
    for c,h in enumerate(hdrs,1):
        cell=ws.cell(row=hr,column=c,value=h); cell.font=h2f; cell.border=tb
    for rn,rd in enumerate(rows_data,hr+1):
        for cn,v in enumerate(rd,1):
            cell=ws.cell(row=rn,column=cn,value=v)
            cell.font=Font(size=9,name="微软雅黑"); cell.alignment=Alignment(vertical='top',wrap_text=True); cell.border=tb
    return hr+1+len(rows_data)+2

# 找到测试数据Sheet的末尾
r = ws2.max_row + 2
r = sec(ws2, r, "3.3 按物料视图-筛选字段与真实数据",
    ["字段","操作符","输入类型","数据中存在值(来自VTable)"],
    [["物料编码","包含/不包含/等于/不等于/为空/不为空","文本输入","12.326.3251/C P0301001/ZJ0301001/TWCP0301001/3001-xxxxx"],
     ["物料名称","同上6种","文本输入","中型胶垫圈/艾玛50系列/艾玛35系列/纸箱/弹簧支架等"],
     ["物料单位","同上6种","文本输入","个"],
     ["仓库","同上6种","文本输入","1F生产仓库/半成品仓/商品-半成品仓/成品仓/铁材仓/塑料仓"],
     ["应发数量","等于/不等于","数字输入","1/4/22/100/39998/999.95等"],
     ["实发数量","等于/不等于","数字输入","-"],
     ["排产开始日期","范围","日期选择器","-"],
     ["排产结束日期","范围","日期选择器","-"]])

fn = f"测试用例_{L1}_{L2}_筛选查询_{date.today().isoformat()}.xlsx"
fp = os.path.join("screenshots", fn)
try:
    wb.save(fp)
    print(f"OK: {os.path.abspath(fp)}")
    print(f"追加了 {len(TC2)} 条按物料视图用例")
    print(f"总计: 原有29条 + 新增{len(TC2)}条 = {29+len(TC2)}条")
except PermissionError:
    wb.save(fn)
    print(f"Fallback: {os.path.abspath(fn)}")
