# -*- coding: utf-8 -*-
# 制令单明细表 测试用例导出脚本
# 数据来源：DrissionPage MCP 真实探测（scan_vtable_columns / get_column_values / run_js 读筛选区）
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ENTERPRISE_PREFIX = "NB"
DEFAULT_AUTHOR    = "Hooplus1ce"
MODULE_NAME       = "生产管理_制令单明细表"
MODULE_LEVEL1     = "生产管理"
MODULE_LEVEL2     = "制令单明细表"
MODULE_PINYIN     = "ZLDMXB"
OUTPUT_DIR        = "."

def cid(g, n):
    return f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_{g}{n:03d}"

wb = Workbook()
ws1 = wb.active
ws1.title = "测试用例"

HEADERS_18 = [
    "用例编号", "用例标题", "级别", "验证点",
    "一级模块", "二级模块", "测试类型", "功能",
    "前置条件", "测试步骤", "测试数据", "预期结果",
    "测试结果", "执行人", "执行时间", "编写人", "编写时间", "备注",
]
ws1.append(HEADERS_18)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

PRIORITY_STYLES = {
    "高级": (PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid"),
           Font(bold=True, color="9C0006", size=10)),
    "中级": (PatternFill(start_color="FFE8D0", end_color="FFE8D0", fill_type="solid"),
           Font(bold=True, color="9C6500", size=10)),
    "低级": (PatternFill(start_color="FFF5C0", end_color="FFF5C0", fill_type="solid"),
           Font(bold=True, color="806000", size=10)),
}

def apply_priority_style(cell):
    s = PRIORITY_STYLES.get(cell.value)
    if s:
        cell.fill, cell.font = s

PRE = ("1. 已登录诺贝科技SCM演示系统(demo17-scm.hoolinks.com)\n"
       "2. 已进入「生产管理-制令单明细表」列表页(iframe: manufactureOrderReport)\n"
       "3. 表格已加载制令单数据(默认92行)")

# 每行 18 字段，顺序对齐 HEADERS_18
test_cases = [
    # ============ F 组：筛选查询 ============
    [cid("F", 1), "按制令单单号包含筛选", "中级", "筛选结果仅含目标单号",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 点击筛选区「展开」按钮展开完整筛选条件\n2. 在「制令单单号包含」输入框输入 MO20260627\n3. 点击查询按钮",
     "制令单单号:MO20260627",
     "表格「制令单单号」列所有值均包含「MO20260627」(如MO202606270046)，无不符合记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "真实单号MO202606270046；可用get_column_values验证列值"],

    [cid("F", 2), "按制令单类型筛选普通制令单", "中级", "筛选结果均为普通制令单",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「制令单类型等于」下拉选择「普通制令单」\n3. 点击查询",
     "制令单类型:普通制令单",
     "表格「制令单类型」列所有值均等于「普通制令单」，不出现包装/返工维修/重组指令单",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "枚举:普通/包装/返工维修/重组指令单"],

    [cid("F", 3), "按制造排产状态筛选未排产", "中级", "筛选结果均为未排产",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「制造排产状态等于」下拉选择「未排产」\n3. 点击查询",
     "制造排产状态:未排产",
     "表格「制造排产状态」列所有值均等于「未排产」",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "枚举:已排产/未排产"],

    [cid("F", 4), "按生产排产状态筛选部分排产", "中级", "筛选结果均为部分排产",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「生产排产状态等于」下拉选择「部分排产」\n3. 点击查询",
     "生产排产状态:部分排产",
     "表格「生产排产状态」列所有值均等于「部分排产」",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "枚举:已排产/待排产/部分排产"],

    [cid("F", 5), "按创建时间范围筛选", "中级", "仅返回区间内创建的制令单",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 「创建时间范围」开始日期选 2026-06-01，结束日期选 2026-06-30\n2. 点击查询",
     "创建时间:2026-06-01 ~ 2026-06-30",
     "表格「创建日期」列所有值均落在 2026-06-01 至 2026-06-30 区间内",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "默认筛选区即含创建时间范围"],

    [cid("F", 6), "按成品编码包含筛选", "中级", "筛选结果均含目标成品编码",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「成品编码包含」输入 2001-0058\n3. 点击查询",
     "成品编码:2001-0058",
     "表格「成品编码」列所有值均包含「2001-0058」(如2001-0058-00)",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "真实编码2001-0058-00"],

    [cid("F", 7), "多条件组合筛选", "中级", "多条件同时生效",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「制令单类型等于」选「包装制令单」\n3. 「制造排产状态等于」选「已排产」\n4. 点击查询",
     "制令单类型:包装制令单;制造排产状态:已排产",
     "表格结果行同时满足:制令单类型=包装制令单 且 制造排产状态=已排产",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "组合筛选AND逻辑"],

    [cid("F", 8), "筛选无结果(不存在的单号)", "低级", "无匹配时表格清空且无报错",
     MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「制令单单号包含」输入 MO9999999999(不存在)\n3. 点击查询",
     "制令单单号:MO9999999999",
     "表格无数据行，显示空表/无数据提示，不报系统错误",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "异常输入边界"],

    [cid("F", 9), "重置筛选条件恢复全量", "低级", "重置后恢复默认全量数据",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录SCM系统\n2. 已进入制令单明细表并执行过筛选(结果非全量)",
     "1. 点击筛选区「重置」按钮\n2. 点击查询",
     "无",
     "筛选条件全部清空，表格恢复默认全量数据(约92行)",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), ""],

    [cid("F", 10), "按数量等于筛选", "低级", "精确数量匹配",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     PRE,
     "1. 展开筛选区\n2. 「数量等于」输入 10990\n3. 点击查询",
     "数量:10990",
     "表格「数量」列所有值均等于 10990",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "真实数量值10990"],

    # ============ I 组：交互(按钮/表格) ============
    [cid("I", 1), "点击新增跳转制令单新增页", "高级", "新增入口可用且正确跳转",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "新增",
     PRE,
     "1. 点击工具栏「新增」按钮",
     "无",
     "新增打开制令单新增页签，iframe URL 切换为 prodctionOrderCreate，新增表单加载完成",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "核心阻塞流程；用get_active_frame验证URL"],

    [cid("I", 2), "未选行点击批量备料提示选择", "中级", "未选行时阻止操作并提示",
     MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "批量备料",
     PRE,
     "1. 不勾选任何表格行\n2. 点击工具栏「批量备料」按钮",
     "无",
     "系统弹出提示消息(如「请选择制令单」)，不进入备料操作",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "异常流程；用detect_modal验证提示"],

    [cid("I", 3), "点击物料查询打开弹窗", "中级", "物料查询入口可用",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "物料查询",
     PRE,
     "1. 勾选1行制令单\n2. 点击工具栏「物料查询」按钮",
     "无",
     "弹出物料查询弹窗/抽屉，展示该制令单的物料明细",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "用detect_modal验证弹窗"],

    [cid("I", 4), "点击流程设置打开弹窗", "中级", "流程设置入口可用",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "流程设置",
     PRE,
     "1. 勾选1行制令单\n2. 点击工具栏「流程设置」按钮",
     "无",
     "弹出流程设置弹窗，展示该制令单的流程配置",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "用detect_modal验证弹窗"],

    [cid("I", 5), "点击导出触发文件下载", "中级", "导出功能可用",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "导出",
     PRE,
     "1. 点击工具栏「导出」按钮",
     "无",
     "浏览器触发文件下载(xls/xlsx)，文件含当前查询结果数据",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "下载行为验证"],

    [cid("I", 6), "表头复选框全选与取消", "低级", "全选勾选所有行",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格操作",
     PRE,
     "1. 点击表头复选框(列0)勾选全选\n2. 再次点击取消全选",
     "无",
     "勾选时所有数据行复选框置为选中；取消时全部清除选中",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "VTable col0 checkbox"],

    [cid("I", 7), "客户编码列升序排序", "中级", "排序按客户编码升序",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "排序",
     PRE,
     "1. 点击表头「客户编码」列的排序图标(首次升序)",
     "无",
     "表格数据按「客户编码」升序排列，序号列重排",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col6 sort_normal图标"],

    [cid("I", 8), "数量列降序排序", "中级", "排序按数量降序",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "排序",
     PRE,
     "1. 点击表头「数量」列排序图标两次(升序→降序)",
     "无",
     "表格数据按「数量」降序排列，最大值10990置顶",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col17 sort_normal；真实最大值10990"],

    [cid("I", 9), "制令单单号列筛选图标弹出筛选菜单", "低级", "列筛选菜单可用",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "筛选",
     PRE,
     "1. 点击表头「制令单单号」列的筛选图标(filter-icon)",
     "无",
     "弹出VTable筛选菜单(.vtable-filter-menu)，含筛选输入与选项",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col3 filter-icon；菜单非ant-dropdown"],

    [cid("I", 10), "点击制令单单号单元格跳转详情", "高级", "单号单元格可跳转详情",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "跳转",
     PRE,
     "1. 点击某行「制令单单号」单元格(如MO202606270046)",
     "制令单单号:MO202606270046",
     "跳转至该制令单详情/编辑页，iframe URL 切换，表单加载该单数据",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col3 popup-candidate；用get_active_frame验证"],

    [cid("I", 11), "点击制造排产状态单元格弹窗", "中级", "排产状态单元格可查看明细",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "弹窗",
     PRE,
     "1. 点击某行「制造排产状态」单元格",
     "无",
     "弹出该制令单的排产状态明细弹窗",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col8 popup-candidate"],

    [cid("I", 12), "点击领料进度单元格弹窗", "低级", "进度单元格可查看明细",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "弹窗",
     PRE,
     "1. 点击某行「领料进度」单元格(如10%)",
     "领料进度:10%",
     "弹出该制令单的领料明细弹窗，展示领料进度构成",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "col20 popup-candidate；真实进度10%"],

    # ============ B 组：批量/导出 ============
    [cid("B", 1), "选中多行后批量备料", "高级", "批量备料核心流程可用",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "批量备料",
     PRE,
     "1. 勾选多行制令单(≥2行)\n2. 点击工具栏「批量备料」按钮",
     "勾选行数:≥2",
     "进入批量备料操作界面/弹窗，展示选中制令单的备料信息",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "核心阻塞流程"],

    [cid("B", 2), "导出当前筛选结果", "中级", "导出反映当前筛选条件",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "导出",
     "1. 已登录SCM系统\n2. 已在制令单明细表执行筛选(结果非全量)",
     "1. 保持筛选条件\n2. 点击工具栏「导出」",
     "无",
     "导出文件仅含当前筛选结果数据，行数与筛选后表格一致",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), ""],

    # ============ P 组：页面级 ============
    [cid("P", 1), "展开与收起筛选区", "低级", "筛选区展开收起正常",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "布局",
     PRE,
     "1. 点击「展开▼」按钮展开筛选区\n2. 再次点击(变收起)收起筛选区",
     "无",
     "展开时显示全部22个筛选字段；收起时仅显示默认创建时间范围，筛选值保留",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "展开按钮文字含▼"],

    [cid("P", 2), "页签切换制令单新增与明细表", "低级", "多页签切换状态保留",
     MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "布局",
     "1. 已登录SCM系统\n2. 已打开「制令单明细表」与「制令单新增」两个页签",
     "1. 点击顶部页签「制令单新增」\n2. 再点击页签「制令单明细表」切回",
     "无",
     "切换页签后各页签状态独立保留，明细表表格数据不丢失",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), ""],
]

for _r in test_cases:
    assert len(_r) == 18, f"字段数!=18: {len(_r)} -> {_r[0] if _r else 'empty'}"
for row_data in test_cases:
    ws1.append(row_data)
    row_idx = ws1.max_row
    for cell_idx, cell in enumerate(ws1[row_idx], 1):
        cell.border = thin_border
        if cell_idx in (2, 4, 9, 10, 11, 12):
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_priority_style(ws1.cell(row=row_idx, column=3))

COL_WIDTHS = [18, 42, 12, 42, 18, 18, 18, 12, 50, 44, 44, 50, 10, 10, 12, 12, 12, 0]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[chr(64 + i)].width = w
ws1.freeze_panes = "E2"
ws1.auto_filter.ref = f"A1:R{ws1.max_row}"

# ===================== Sheet 2: 测试数据 =====================
ws2 = wb.create_sheet(title="测试数据")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11, name="微软雅黑")
header2_font = Font(bold=True, size=10, name="微软雅黑")

def write_section(ws, start_row, title, headers, data_rows):
    col_count = len(headers)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = section_font
    title_cell.fill = section_fill
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    for c in range(1, col_count + 1):
        ws.cell(row=start_row, column=c).fill = section_fill
        ws.cell(row=start_row, column=c).border = thin_border
    hdr_row = start_row + 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font = header2_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_offset, row_data in enumerate(data_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=hdr_row + r_offset, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    return hdr_row + 1 + len(data_rows) + 2

# 3.1 测试数据配置（用例概览）
overview = []
for i, tc in enumerate(test_cases, 1):
    overview.append([i, tc[0], tc[4], tc[5], tc[7], tc[1], tc[2], tc[6], tc[15], tc[16]])
row = 1
row = write_section(ws2, row, "3.1 测试数据配置",
    ["序号", "系统编号", "一级模块", "二级模块", "功能", "用例标题", "级别", "测试类型", "编写人", "编写日期"], overview)

# 3.2 筛选字段说明（真实探测）
filter_fields = [
    ["生产排产状态", "下拉等于", "已排产/待排产/部分排产", "枚举3值"],
    ["审批状态", "下拉等于", "[待确认]图标渲染取值失败", "表格列纯图标渲染，get_column_values读不到"],
    ["创建时间范围", "日期范围", "2026-06-01~2026-06-30", "默认显示的筛选条件"],
    ["制令单单号", "输入包含", "MO202606270046", "格式MO+8位日期+4位序号"],
    ["销售单号", "输入包含", "-", "popup-candidate列"],
    ["客户编码", "下拉包含", "-", ""],
    ["客户名称", "下拉包含", "-", ""],
    ["制造排产状态", "下拉等于", "已排产/未排产", "枚举2值"],
    ["成品编码", "输入包含", "2001-0058-00", ""],
    ["成品名称", "输入包含", "-", ""],
    ["成品型号", "输入包含", "-", ""],
    ["成品单位", "输入包含", "-", ""],
    ["生产部门", "下拉包含", "-", ""],
    ["数量", "输入等于", "10990", "数值"],
    ["已缴库数量", "输入等于", "-", "数值"],
    ["未完工量", "输入等于", "-", "数值"],
    ["领料时间范围", "日期范围", "-", ""],
    ["发料时间范围", "日期范围", "-", ""],
    ["制造部门", "输入包含", "-", ""],
    ["新建订单(OA)", "下拉+勾选", "-", "禁用状态disabled"],
    ["订单修改:关联OA流程:制令单变更", "下拉", "-", "禁用状态disabled"],
    ["关键词", "输入", "可使用空格分隔多个关键词", "通用关键词搜索"],
]
row = write_section(ws2, row, "3.2 筛选字段说明",
    ["字段名称", "输入方式", "对应测试数据", "说明"], filter_fields)

# 3.3 工具栏按钮一览表
buttons = [
    ["制令单明细表", "新增", "否"],
    ["制令单明细表", "导出", "否(导出当前结果)"],
    ["制令单明细表", "物料查询", "是(需选行)"],
    ["制令单明细表", "批量备料", "是(需选行)"],
    ["制令单明细表", "流程设置", "是(需选行)"],
]
row = write_section(ws2, row, "3.3 各页签工具栏按钮一览表",
    ["状态页签", "按钮名称", "是否需要勾选行"], buttons)

# 3.4 VTable列定义一览表（真实 scan_vtable_columns 输出）
vtable_cols = [
    ["复选框", "_vtable_checkbox", "勾选状态", "全选复选框(col0)"],
    ["序号", "序号", "数字", "下拉菜单dropdownIcon"],
    ["创建日期", "创建日期", "日期时间", "筛选filter"],
    ["制令单单号", "制令单单号", "MO+8位日期+4位序号", "筛选filter；body可popup跳转"],
    ["销售单号", "销售单号", "文本", "筛选filter；body可popup"],
    ["审批状态", "审批状态", "图标渲染(取值失败)", "筛选filter；视觉/原始值均读不到"],
    ["客户编码", "客户编码", "XXXX-XXXX-XX", "筛选+排序sort_normal"],
    ["客户名称", "客户名称", "中文", "筛选+排序"],
    ["制造排产状态", "制造排产状态", "已排产/未排产", "筛选+排序；body可popup"],
    ["生产排产状态", "生产排产状态", "已排产/待排产/部分排产", "筛选+排序；body可popup"],
    ["制令单类型", "制令单类型", "普通/包装/返工维修/重组指令", "筛选+排序；body可popup"],
    ["成品编码", "成品编码", "XXXX-XXXX-XX", "筛选+排序；body可popup"],
    ["成品规格", "成品规格", "文本", "筛选+排序"],
    ["成品名称", "成品名称", "中文", "筛选+排序"],
    ["成品型号", "成品型号", "文本", "筛选filter"],
    ["成品单位", "成品单位", "文本", "筛选filter"],
    ["制造部门", "制造部门", "中文", "筛选+排序"],
    ["数量", "数量", "数值(如10990)", "筛选+排序"],
    ["已缴库量", "已缴库量", "数值", "筛选+排序"],
    ["未完工量", "未完工量", "数值", "筛选+排序"],
    ["领料进度", "领料进度", "百分比(如10%)", "筛选filter；body可popup"],
    ["缴库进度", "缴库进度", "百分比", "筛选filter；body可popup"],
    ["发料进度", "发料进度", "百分比", "筛选filter；body可popup"],
    ["领料时间", "领料时间", "日期时间", "筛选filter"],
    ["发料时间", "发料时间", "日期时间", "筛选filter"],
]
row = write_section(ws2, row, "3.4 VTable列定义一览表",
    ["列标题", "字段名", "数据格式说明", "列头交互能力"], vtable_cols)

# ===================== 保存 =====================
filename = f"测试用例_{MODULE_NAME}_{date.today().isoformat()}.xlsx"
os.makedirs(OUTPUT_DIR, exist_ok=True)
filepath = os.path.join(OUTPUT_DIR, filename)
try:
    wb.save(filepath)
    print(f"OK 已生成: {os.path.abspath(filepath)}")
    print(f"用例数: {len(test_cases)}")
except PermissionError:
    fallback = os.path.join(".", filename)
    wb.save(fallback)
    print(f"降级保存至: {os.path.abspath(fallback)}")
