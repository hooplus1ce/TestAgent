#!/usr/bin/env python3
"""Export test cases for 制令单明细表 to Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date

ENTERPRISE_PREFIX = "NB"
MODULE_PINYIN = "ZLDMB"
DEFAULT_AUTHOR = "Hooplus1ce"
TODAY = date.today().isoformat()
OUTPUT_FILE = f"测试用例_{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_{TODAY}.xlsx"

MODULE_L1 = "生产管理"
MODULE_L2 = "制令单明细表"

def tc(id_suffix, title, priority, verify_point, test_type, func, precond, steps, test_data, expected):
    return [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_{id_suffix}",
        title, priority, verify_point,
        MODULE_L1, MODULE_L2, test_type, func,
        precond, steps, test_data, expected,
        "", "", "", DEFAULT_AUTHOR, TODAY, ""
    ]

test_cases = []

# ==================== F Series (筛选查询) ====================

test_cases.append(tc("F001", "视图切换—点击「按成品」查看成品维度数据", "P0",
    "点击「按成品」后，VTable切换为成品维度视图，展示制令单的成品信息",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 若当前为「按物料」视图，点击「按成品」选项卡",
    "",
    "1. 「按成品」选项卡高亮\n2. VTable表头包含成品编码、成品名称、成品规格等成品相关列\n3. 表格展示90条制令单记录"))

test_cases.append(tc("F002", "视图切换—点击「按物料」查看物料维度数据", "P1",
    "点击「按物料」后，VTable切换为物料维度视图，展示制令单的物料信息",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击「按物料」选项卡",
    "",
    "1. 「按物料」选项卡高亮\n2. VTable表头包含物料编码、物料规格、物料名称、仓库、应发数量、实发数量等物料相关列\n3. 数据内容与物料维度一致"))

test_cases.append(tc("F003", "筛选—生产排产状态等于「已排产」", "P0",
    "筛选生产排产状态=已排产后，VTable仅显示已排产的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「生产排产状态」字段选择「等于」\n2. 在值输入框下拉选择「已排产」\n3. 点击查询",
    "生产排产状态=已排产",
    "筛选后表格仅显示生产排产状态为「已排产」的制令单（约7条），无状态非「已排产」的记录"))

test_cases.append(tc("F004", "筛选—生产排产状态等于「待排产」", "P0",
    "筛选生产排产状态=待排产后，VTable仅显示待排产的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「生产排产状态」字段选择「等于」\n2. 在值输入框下拉选择「待排产」\n3. 点击查询",
    "生产排产状态=待排产",
    "筛选后表格仅显示生产排产状态为「待排产」的制令单（约75条），无状态非「待排产」的记录"))

test_cases.append(tc("F005", "筛选—生产排产状态等于「部分排产」", "P0",
    "筛选生产排产状态=部分排产后，VTable仅显示部分排产的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「生产排产状态」字段选择「等于」\n2. 在值输入框下拉选择「部分排产」\n3. 点击查询",
    "生产排产状态=部分排产",
    "筛选后表格仅显示生产排产状态为「部分排产」的制令单（约8条），无状态非「部分排产」的记录"))

test_cases.append(tc("F006", "筛选—制造排产状态等于「已排产」", "P0",
    "筛选制造排产状态=已排产后，VTable仅显示已排产的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制造排产状态」字段选择「等于」\n2. 在值输入框下拉选择「已排产」\n3. 点击查询",
    "制造排产状态=已排产",
    "筛选后表格仅显示制造排产状态为「已排产」的制令单，无状态非「已排产」的记录"))

test_cases.append(tc("F007", "筛选—制造排产状态等于「未排产」", "P0",
    "筛选制造排产状态=未排产后，VTable仅显示未排产的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制造排产状态」字段选择「等于」\n2. 在值输入框下拉选择「未排产」\n3. 点击查询",
    "制造排产状态=未排产",
    "筛选后表格仅显示制造排产状态为「未排产」的制令单，无状态非「未排产」的记录"))

test_cases.append(tc("F008", "筛选—制令单单号包含指定单号", "P0",
    "筛选制令单单号包含输入值后，VTable仅显示单号包含该值的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制令单单号」字段输入「MO202606」\n2. 点击查询",
    "制令单单号包含MO202606",
    "筛选后所有制令单单号均包含「MO202606」，无不符合条件的记录"))

test_cases.append(tc("F009", "筛选—销售单号包含指定单号", "P1",
    "筛选销售单号包含输入值后，VTable仅显示销售单号包含该值的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「销售单号」字段输入「SO202606」\n2. 点击查询",
    "销售单号包含SO202606",
    "筛选后所有销售单号均包含「SO202606」，无不符合条件的记录"))

test_cases.append(tc("F010", "筛选—客户编码包含指定编码", "P1",
    "筛选客户编码包含输入值后，VTable仅显示该客户的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「客户编码」字段输入「CUS202602」\n2. 点击查询",
    "客户编码包含CUS202602",
    "筛选后所有客户编码均包含「CUS202602」，无不符合条件的记录"))

test_cases.append(tc("F011", "筛选—客户名称包含指定名称", "P1",
    "筛选客户名称包含输入值后，VTable仅显示该客户的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「客户名称」字段输入「诺贝科技」\n2. 点击查询",
    "客户名称包含诺贝科技",
    "筛选后所有客户名称均包含「诺贝科技」，无不符合条件的记录"))

test_cases.append(tc("F012", "筛选—成品编码包含指定编码", "P1",
    "筛选成品编码包含输入值后，VTable仅显示该成品的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「成品编码」字段输入「2001-0058」\n2. 点击查询",
    "成品编码包含2001-0058",
    "筛选后所有成品编码均包含「2001-0058」，无不符合条件的记录"))

test_cases.append(tc("F013", "筛选—成品名称包含指定名称", "P1",
    "筛选成品名称包含输入值后，VTable仅显示该成品的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「成品名称」字段输入「艾玛35」\n2. 点击查询",
    "成品名称包含艾玛35",
    "筛选后所有成品名称均包含「艾玛35」，无不符合条件的记录"))

test_cases.append(tc("F014", "筛选—成品单位等于指定单位", "P2",
    "筛选成品单位包含输入值后，VTable仅显示该单位的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「成品单位」字段输入「个」\n2. 点击查询",
    "成品单位=个",
    "筛选后所有成品单位均为「个」，无不符合条件的记录"))

test_cases.append(tc("F015", "筛选—生产部门包含指定部门", "P1",
    "筛选生产部门包含输入值后，VTable仅显示该部门的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「生产部门」字段输入「安装车间」\n2. 点击查询",
    "生产部门包含安装车间",
    "筛选后所有生产部门均为「安装车间」，无其他部门的记录"))

test_cases.append(tc("F016", "筛选—创建时间范围查询", "P0",
    "筛选创建时间在指定范围内后，VTable仅显示该时间范围的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「创建时间」字段设置开始日期为2026-06-01\n2. 设置结束日期为2026-06-30\n3. 点击查询",
    "创建时间=2026-06-01~2026-06-30",
    "筛选后所有记录的创建日期均在2026-06-01至2026-06-30范围内，无超出范围的记录"))

test_cases.append(tc("F017", "筛选—数量等于指定值", "P1",
    "筛选数量等于输入值后，VTable仅显示该数量的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「数量」字段输入「3」\n2. 点击查询",
    "数量=3",
    "筛选后所有制令单的数量均为3，无数值不为3的记录"))

test_cases.append(tc("F018", "筛选—已缴库数量等于0", "P1",
    "筛选已缴库数量等于0后，VTable仅显示未缴库的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「已缴库数量」字段输入「0」\n2. 点击查询",
    "已缴库数量=0",
    "筛选后所有制令单的已缴库数量均为0，无已缴库的记录"))

test_cases.append(tc("F019", "筛选—未完工量等于指定值", "P1",
    "筛选未完工量等于输入值后，VTable仅显示匹配的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「未完工量」字段输入「3」\n2. 点击查询",
    "未完工量=3",
    "筛选后所有制令单的未完工量均为3，无数值不为3的记录"))

test_cases.append(tc("F020", "筛选—领料时间范围查询", "P1",
    "筛选领料时间在指定范围内后，VTable仅显示该时间范围的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「领料时间」字段设置日期范围为2026-06-01至2026-06-30\n2. 点击查询",
    "领料时间=2026-06-01~2026-06-30",
    "筛选后所有领料时间均在指定范围内，无超出范围的记录"))

test_cases.append(tc("F021", "组合筛选—生产排产状态等于待排产且成品编码包含指定值", "P0",
    "组合多个筛选条件后，VTable仅显示同时满足所有条件的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「生产排产状态」选择「等于」→「待排产」\n2. 在「成品编码」字段输入「2001」\n3. 点击查询",
    "生产排产状态=待排产, 成品编码包含2001",
    "筛选后所有行生产排产状态均为「待排产」且成品编码包含「2001」，同时满足两个条件"))

test_cases.append(tc("F022", "组合筛选—制造排产状态+生产部门", "P0",
    "使用多个字段组合筛选，验证多条件联合过滤的正确性",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制造排产状态」选择「等于」→「未排产」\n2. 在「生产部门」字段输入「安装车间」\n3. 点击查询",
    "制造排产状态=未排产, 生产部门包含安装车间",
    "筛选后所有行制造排产状态均为「未排产」且生产部门为「安装车间」"))

test_cases.append(tc("F023", "筛选—点击重置按钮清空所有筛选条件", "P0",
    "点击重置后，所有筛选字段恢复到默认值，VTable显示全部数据",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区设置任意筛选条件\n2. 点击「重置」按钮",
    "",
    "1. 所有筛选字段清空为默认值\n2. 表格恢复显示全部90条制令单记录"))

test_cases.append(tc("F024", "筛选—审批状态等于指定状态", "P1",
    "筛选审批状态等于指定值后，VTable仅显示该状态的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「审批状态」字段选择「等于」\n2. 在下拉中选择一个审批状态\n3. 点击查询",
    "审批状态=指定值",
    "筛选后表格仅显示指定审批状态的制令单，无不符合条件的记录"))

test_cases.append(tc("F025", "筛选—成品型号包含指定型号", "P2",
    "筛选成品型号包含输入值后，VTable仅显示该型号的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「成品型号」字段输入指定型号\n2. 点击查询",
    "成品型号包含指定值",
    "筛选后所有成品型号均包含输入值，无不符合条件的记录"))

test_cases.append(tc("F026", "筛选—制造部门包含指定部门", "P1",
    "筛选制造部门包含输入值后，VTable仅显示该部门的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制造部门」字段输入「冲压车间」\n2. 点击查询",
    "制造部门包含冲压车间",
    "筛选后所有制造部门均为「冲压车间」，无其他部门的记录"))

test_cases.append(tc("F027", "筛选—发料时间范围查询", "P2",
    "筛选发料时间在指定范围内后，VTable仅显示该时间范围的制令单",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「发料时间」字段设置日期范围\n2. 点击查询",
    "发料时间=指定范围",
    "筛选后所有发料时间在指定范围内，无超出范围的记录"))

test_cases.append(tc("F028", "筛选—筛选条件清空后查询全部数据", "P1",
    "清空所有筛选条件并查询后，VTable恢复到显示全部90条记录",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区输入任意筛选条件并查询\n2. 清空所有筛选条件\n3. 再次点击查询",
    "",
    "筛选条件清空并查询后，表格恢复到显示全部90条制令单记录"))

test_cases.append(tc("F029", "筛选—制令单单号精确匹配查询", "P1",
    "使用制令单单号精确搜索，验证精确匹配功能",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制令单单号」字段输入「MO202606270046」\n2. 点击查询",
    "制令单单号=MO202606270046",
    "筛选后表格仅显示单号为MO202606270046的制令单，其他制令单不显示"))

test_cases.append(tc("F030", "筛选—使用不存在的条件查询无匹配结果", "P2",
    "使用不可能匹配的条件筛选时，表格显示为空",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在筛选区「制令单单号」字段输入一个不存在的单号「XXXXXXXXX」\n2. 点击查询",
    "制令单单号=XXXXXXXXX",
    "筛选后表格显示无数据（空状态），页面不报错"))

# ==================== I Series (交互操作) ====================

test_cases.append(tc("I001", "新增—点击新增按钮跳转至制令单新增页面", "P0",
    "点击新增按钮后，页面跳转到制令单新增Tab页",
    "功能测试", "新增",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击工具栏「新增」按钮",
    "",
    "1. 新增一个「制令单新增」Tab页签\n2. 当前页面切换为制令单新增表单页面\n3. Tab栏数量变为2个\n4. 新Tab的iframe加载制令单新增表单"))

test_cases.append(tc("I002", "新增—关闭新增Tab页返回制令单明细表", "P1",
    "关闭制令单新增Tab后，页面返回制令单明细表",
    "功能测试", "新增",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面",
    "1. 点击「新增」跳转到制令单新增\n2. 点击新增Tab页签上的关闭×按钮",
    "",
    "1. 制令单新增Tab关闭\n2. 页面回到制令单明细表Tab\n3. Tab栏数量恢复为1个\n4. 制令单明细表数据正常显示"))

test_cases.append(tc("I003", "物料查询—未勾选制令单时点击提示", "P1",
    "未勾选任何制令单时点击物料查询，弹出提示信息",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载\n4. 未勾选任何行",
    "1. 点击工具栏「物料查询」按钮",
    "",
    "1. 弹出提示信息「请勾选一项」\n2. 页面不跳转\n3. 提示信息可手动关闭"))

test_cases.append(tc("I004", "批量备料—未选择制令单时点击提示", "P1",
    "未选择制令单时点击批量备料，弹出提示信息",
    "功能测试", "批量操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载\n4. 未勾选任何行",
    "1. 点击工具栏「批量备料」按钮",
    "",
    "1. 弹出提示信息「请先选择制令单」\n2. 页面不跳转\n3. 提示信息可手动关闭"))

test_cases.append(tc("I005", "导出—点击导出按钮触发下载", "P1",
    "点击导出按钮后，触发文件下载或导出功能",
    "功能测试", "导出",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击工具栏「导出」按钮",
    "",
    "1. 触发导出/下载操作\n2. 页面不报错"))

test_cases.append(tc("I006", "流程设置—点击流程设置按钮", "P2",
    "点击流程设置按钮后，检查页面响应",
    "功能测试", "设置",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击工具栏「流程设置」按钮",
    "",
    "页面响应不报错"))

test_cases.append(tc("I007", "页面—点击收起/展开切换筛选区可见性", "P2",
    "点击收起按钮折叠筛选区，点击展开按钮重新显示筛选区",
    "功能测试", "页面布局",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 筛选区已展开（按钮显示「收起▲」）",
    "1. 点击「收起▲」按钮\n2. 观察筛选区状态\n3. 再次点击「展开▼」按钮",
    "",
    "1. 点击收起▲后，筛选区折叠隐藏\n2. 按钮文字变为「展开▼」\n3. 点击展开▼后，筛选区重新显示\n4. 按钮文字恢复为「收起▲」"))

test_cases.append(tc("I008", "表格—单击复选框选中/取消选中行", "P0",
    "单击行首复选框可切换该行的选中状态",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 将鼠标移动到第一行数据行的复选框区域\n2. 单击复选框\n3. 再次单击同一复选框",
    "",
    "1. 单击后该行高亮，复选框显示选中状态\n2. 再次单击后取消选中\n3. 选中状态与视觉反馈一致"))

test_cases.append(tc("I009", "表格—点击数量列表头排序图标切换排序", "P1",
    "点击数量列表头排序图标，数据按数量升序/降序排列",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「数量」列的表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标从↕切换为↑（升序）\n2. 数据按数量从小到大排列\n3. 再次点击切换为↓（降序）"))

test_cases.append(tc("I010", "表格—点击制令单单号列排序", "P1",
    "点击制令单单号列表头排序图标，数据按单号排序",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「制令单单号」列的表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标状态变化（↕→↑或↓）\n2. 数据按单号重新排列"))

test_cases.append(tc("I011", "表格—点击已缴库量表头排序图标", "P2",
    "点击已缴库量排序图标，数据按数值排序",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「已缴库量」列的表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标状态变化\n2. 数据按已缴库量重新排列"))

test_cases.append(tc("I012", "表格—点击未完工量表头排序图标", "P2",
    "点击未完工量排序图标，数据按数值排序",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「未完工量」列的表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标状态变化\n2. 数据按未完工量重新排列"))

test_cases.append(tc("I013", "Tab切换—切换到流程发起规则设置", "P1",
    "点击流程发起规则设置Tab，页面切换到规则配置视图",
    "功能测试", "页面切换",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击「流程发起规则设置」Tab",
    "",
    "1. 页面切换为流程发起规则配置界面\n2. 原制令单表格隐藏\n3. Tab高亮切换"))

test_cases.append(tc("I014", "Tab切换—切换到流程设计", "P1",
    "点击流程设计Tab，页面切换到流程设计视图",
    "功能测试", "页面切换",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击「流程设计」Tab",
    "",
    "1. 页面切换为流程设计界面\n2. 原制令单表格隐藏\n3. Tab高亮切换"))

test_cases.append(tc("I015", "视图切换—按成品与按物料来回切换", "P1",
    "在按成品和按物料视图之间来回切换，数据正确刷新",
    "功能测试", "视图切换",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 点击「按物料」切换到物料视图\n2. 观察表格列变化\n3. 点击「按成品」切换回成品视图",
    "",
    "1. 切换到按物料后，表头出现物料相关列（物料编码、物料名称、仓库等）\n2. 切换回按成品后，表头恢复为成品相关列\n3. 切换过程不报错"))

test_cases.append(tc("I016", "表格—拖动列表头调整列宽", "P2",
    "拖动列头边框可调整列宽",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 将鼠标移动到两列表头交界处\n2. 光标变为拖拽指针后长按左键拖动\n3. 松开鼠标",
    "",
    "1. 列宽跟随拖动方向变化\n2. 列内容显示正常无截断\n3. 其他列布局自适应"))

test_cases.append(tc("I017", "表格—双击制令单单号列查看详情", "P0",
    "双击制令单单号链接列后，弹出详情或跳转",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在制令单单号列找到一条制令单号\n2. 双击该单元格",
    "",
    "1. 触发弹窗显示制令单详情或跳转到详情页面\n2. 弹窗/页面中的关键信息与所选行一致"))

test_cases.append(tc("I018", "表格—双击生产排产状态查看详情", "P1",
    "双击生产排产状态列后，可能弹出排产明细",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到一条生产排产状态为「已排产」的记录\n2. 双击该状态单元格",
    "",
    "1. 触发弹窗或跳转，显示该制令单的排产明细信息\n2. 弹窗中数据与所选行一致"))

test_cases.append(tc("I019", "表格—点击制令单单号列表头筛选图标", "P1",
    "点击制令单单号列表头筛选图标后弹出筛选面板",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「制令单单号」列表头的筛选图标（漏斗形）\n2. 单击筛选图标",
    "",
    "1. 弹出VTable列头筛选面板\n2. 筛选面板包含搜索框和条件选项\n3. 可关闭筛选面板"))

test_cases.append(tc("I020", "表格—点击成品编码列头筛选图标", "P1",
    "点击成品编码列头筛选图标后弹出筛选面板",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「成品编码」列表头的筛选图标\n2. 单击筛选图标",
    "",
    "1. 弹出VTable列头筛选面板\n2. 可输入条件进行列内筛选"))

test_cases.append(tc("I021", "筛选—筛选后重置再重新筛选", "P1",
    "筛选后重置再重新筛选，验证筛选状态完全重置",
    "功能测试", "查询",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 设置筛选条件并查询\n2. 点击重置清空条件\n3. 设置新的筛选条件\n4. 再次查询",
    "",
    "1. 第一次筛选结果正确\n2. 重置后表格恢复到全部数据\n3. 第二次筛选使用新条件后结果正确"))

test_cases.append(tc("I022", "表格—水平滚动查看隐藏列", "P2",
    "水平滚动VTable可查看右侧隐藏的列",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在VTable底部水平滚动条上向右拖动\n2. 观察右侧列的显示情况",
    "",
    "1. 滚动后右侧隐藏列（领料时间、发料时间等）进入视口\n2. 左侧列移出视口\n3. 滚动流畅无卡顿"))

test_cases.append(tc("I023", "表格—垂直滚动查看更多数据行", "P2",
    "垂直滚动VTable可加载更多数据行",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 在VTable垂直滚动条上向下拖动\n2. 观察后续数据行的加载情况",
    "",
    "1. 滚动后显示更多数据行\n2. 共90条记录可全部浏览\n3. 滚动流畅"))

test_cases.append(tc("I024", "表格—点击客户编码排序", "P2",
    "点击客户编码排序图标，数据按编码排序",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「客户编码」列表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标状态变化\n2. 数据按客户编码重新排列"))

test_cases.append(tc("I025", "表格—点击制造部门排序", "P2",
    "点击制造部门排序图标，数据按部门排序",
    "功能测试", "表格操作",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 找到「制造部门」列表头排序图标\n2. 单击排序图标",
    "",
    "1. 排序图标状态变化\n2. 数据按制造部门重新排列"))

# ==================== P Series (页面级) ====================

test_cases.append(tc("P001", "页面—制令单明细表页签显示与激活", "P0",
    "左侧菜单点击制令单明细表后，页签正确显示并激活",
    "功能测试", "页面导航",
    "1. 已登录SCM系统",
    "1. 在左侧菜单栏找到「制令单明细表」\n2. 点击菜单项",
    "",
    "1. 页面顶部出现「制令单明细表」Tab页签\n2. Tab处于激活状态（高亮）\n3. iframe中加载制令单明细表页面\n4. VTable加载并显示90条数据"))

test_cases.append(tc("P002", "页面—制令单明细表页面元素完整性", "P1",
    "制令单明细表页面包含所有核心功能区域",
    "兼容性测试", "页面布局",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面",
    "1. 观察页面的整体布局",
    "",
    "1. 页面包含以下区域：视图切换（按成品/按物料）、筛选区、工具栏（新增/导出/物料查询/批量备料/流程设置/重置）、VTable\n2. 所有区域正常显示"))

test_cases.append(tc("P003", "页面—刷新浏览器后制令单明细表页签状态", "P1",
    "刷新浏览器后，制令单明细表Tab和表格数据重新加载",
    "功能测试", "页面导航",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 表格数据已加载",
    "1. 刷新当前浏览器页面\n2. 等待页面重新加载",
    "",
    "1. 页面重新加载后，制令单明细表Tab仍然激活\n2. VTable重新加载数据\n3. 数据与刷新前一致"))

test_cases.append(tc("P004", "页面—筛选区折叠后表格区域自适应", "P2",
    "筛选区折叠后，VTable区域占据更多空间",
    "功能测试", "页面布局",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面\n3. 筛选区已展开",
    "1. 点击「收起▲」折叠筛选区\n2. 观察VTable区域的高度变化",
    "",
    "1. 筛选区折叠隐藏\n2. VTable可视区域高度增加\n3. 表格数据仍然完整显示"))

test_cases.append(tc("P005", "页面—制令单明细表与新增Tab之间切换", "P1",
    "制令单明细表和制令单新增两个Tab间切换正常",
    "功能测试", "页面导航",
    "1. 已登录SCM系统\n2. 处于制令单明细表页面",
    "1. 点击「新增」按钮打开制令单新增Tab\n2. 点击「制令单明细表」Tab切回\n3. 再次点击「制令单新增」Tab",
    "",
    "1. 切换到制令单明细表Tab时，表格数据正常显示\n2. 切换到制令单新增Tab时，表单页面正常显示\n3. 来回切换不报错"))

# ==================== Verification ====================
ids = [tc[0] for tc in test_cases]
assert len(ids) == len(set(ids)), f"Duplicate IDs: {[id for id in ids if ids.count(id) > 1]}"

print(f"Total test cases: {len(test_cases)}")
print(f"Groups: F={sum(1 for t in test_cases if '_F' in t[0])}, I={sum(1 for t in test_cases if '_I' in t[0])}, P={sum(1 for t in test_cases if '_P' in t[0])}")
print(f"P0={sum(1 for t in test_cases if t[2]=='P0')}, P1={sum(1 for t in test_cases if t[2]=='P1')}, P2={sum(1 for t in test_cases if t[2]=='P2')}")

# ==================== Excel Export ====================
wb = Workbook()
ws = wb.active
ws.title = "测试用例"

# Title
ws.merge_cells('A1:R1')
c = ws['A1']
c.value = f"测试用例 - {MODULE_L1} - {MODULE_L2}"
c.font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
c.alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:R2')
c = ws['A2']
c.value = f"企业: {ENTERPRISE_PREFIX} | 编写人: {DEFAULT_AUTHOR} | 日期: {TODAY} | 用例数: {len(test_cases)}"
c.font = Font(name='微软雅黑', size=10, color='666666')
c.alignment = Alignment(horizontal='center', vertical='center')

# Headers
headers = ['A:用例编号', 'B:用例标题', 'C:级别', 'D:验证点',
           'E:一级模块', 'F:二级模块', 'G:测试类型', 'H:功能',
           'I:前置条件', 'J:测试步骤', 'K:测试数据', 'L:预期结果',
           'M:测试结果', 'N:执行人', 'O:执行时间', 'P:编写人', 'Q:编写时间', 'R:备注']

header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'))

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Priority colors
p0_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
p1_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
p2_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

left_cols = {2, 4, 9, 10, 11, 12}

for row_idx, tc_row in enumerate(test_cases, 4):
    level = tc_row[2]
    if level == 'P0': row_fill = p0_fill
    elif level == 'P1': row_fill = p1_fill
    else: row_fill = p2_fill

    for col_idx, value in enumerate(tc_row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=9, bold=(level == 'P0'))
        cell.fill = row_fill
        cell.border = thin_border
        if col_idx in left_cols:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
col_widths = {'A': 18, 'B': 42, 'C': 10, 'D': 42, 'E': 12, 'F': 16, 'G': 12, 'H': 12,
              'I': 42, 'J': 48, 'K': 28, 'L': 48, 'M': 10, 'N': 12, 'O': 14, 'P': 12, 'Q': 14, 'R': 10}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 28
for r in range(4, 4 + len(test_cases)):
    ws.row_dimensions[r].height = 60

ws.freeze_panes = 'A4'

# ==================== VTable 列定义一览表 ====================
ws2 = wb.create_sheet("VTable列定义一览表")

ws2.merge_cells('A1:H1')
c = ws2['A1']
c.value = f"{MODULE_L1} - {MODULE_L2} VTable 列定义一览表"
c.font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
c.alignment = Alignment(horizontal='center', vertical='center')

vheaders = ['列号', '字段名', '标题', '类型', '宽度', 'Body行为', 'Body详情', '表头图标']
for col_idx, h in enumerate(vheaders, 1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = Font(name='微软雅黑', size=10, bold=True)
    cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

vtable_data = [
    (0,'_vtable_checkbox','','复选框',40,'复选框','可勾选','复选框图标'),
    (1,'index','序号','整数',80,'none','文本','下拉菜单'),
    (2,'createDate','创建日期','日期',160,'none','文本','筛选'),
    (3,'manufactureNo','制令单单号','文本',120,'链接/按钮','自定义弹窗','筛选'),
    (4,'associatedOrderNo','销售单号','文本',120,'链接/按钮','自定义弹窗','筛选'),
    (5,'approvalStatusText','审批状态','文本',100,'none','文本','筛选'),
    (6,'customerCode','客户编码','文本',120,'none','文本','筛选+排序'),
    (7,'customerName','客户名称','文本',120,'none','文本','筛选+排序'),
    (8,'deviceSchedulingStatus','制造排产状态','文本',120,'链接/按钮','自定义弹窗','筛选+排序'),
    (9,'productionSchedulingStatus','生产排产状态','文本',120,'链接/按钮','自定义弹窗','筛选+排序'),
    (10,'orderType','制令单类型','文本',120,'链接/按钮','自定义弹窗','筛选+排序'),
    (11,'goodsCode','成品编码','文本',150,'链接/按钮','自定义弹窗','筛选+排序'),
    (12,'goodsModel','成品规格','文本',120,'none','文本','筛选+排序'),
    (13,'goodsName','成品名称','文本',100,'none','文本','筛选+排序'),
    (14,'goodsModel','成品型号','文本',100,'none','文本','筛选'),
    (15,'goodsUnit','成品单位','文本',60,'none','文本','筛选'),
    (16,'manufactureDepartmentType','制造部门','文本',100,'none','文本','筛选+排序'),
    (17,'productionQty','数量','整数',80,'none','文本','筛选+排序'),
    (18,'payWarehouseQty','已缴库量','整数',80,'none','文本','筛选+排序'),
    (19,'notFinishQty','未完工量','整数',100,'none','文本','筛选+排序'),
    (20,'pickingProgress','领料进度','百分比',120,'链接/按钮','自定义弹窗','筛选'),
    (21,'inWarehouseProgress','缴库进度','百分比',120,'链接/按钮','自定义弹窗','筛选'),
    (22,'issueProgress','发料进度','百分比',120,'链接/按钮','自定义弹窗','筛选'),
    (23,'receiveTime','领料时间','日期',150,'none','文本','筛选'),
    (24,'sendTime','发料时间','日期',150,'none','文本','筛选'),
]

for row_idx, col_data in enumerate(vtable_data, 3):
    for col_idx, value in enumerate(col_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

vt_widths = {'A':8,'B':30,'C':18,'D':10,'E':8,'F':14,'G':16,'H':18}
for col, w in vt_widths.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A3'

# ==================== 汇总 ====================
ws3 = wb.create_sheet("汇总")
summary = [
    ("项目", "内容"),
    ("企业", ENTERPRISE_PREFIX),
    ("领域", "MOM/SCM"),
    ("一级模块", MODULE_L1),
    ("二级模块", MODULE_L2),
    ("测试用例总数", len(test_cases)),
    ("",""),
    ("F - 筛选查询", sum(1 for t in test_cases if '_F' in t[0])),
    ("I - 交互操作", sum(1 for t in test_cases if '_I' in t[0])),
    ("P - 页面级", sum(1 for t in test_cases if '_P' in t[0])),
    ("",""),
    ("P0 - 阻塞", sum(1 for t in test_cases if t[2]=='P0')),
    ("P1 - 重要", sum(1 for t in test_cases if t[2]=='P1')),
    ("P2 - 一般", sum(1 for t in test_cases if t[2]=='P2')),
    ("",""),
    ("编写人", DEFAULT_AUTHOR),
    ("编写日期", TODAY),
]

for row_idx, (label, value) in enumerate(summary, 1):
    c1 = ws3.cell(row=row_idx, column=1, value=label)
    c2 = ws3.cell(row=row_idx, column=2, value=value)
    c1.font = Font(name='微软雅黑', size=10, bold=True)
    c2.font = Font(name='微软雅黑', size=10)
    if row_idx == 1:
        for c in [c1, c2]:
            c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            c.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    c1.alignment = Alignment(horizontal='left', vertical='center')
    c2.alignment = Alignment(horizontal='center', vertical='center')
    for c in [c1, c2]:
        c.border = thin_border

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 20

# Save
wb.save(OUTPUT_FILE)
print(f"\n✓ Excel exported: {OUTPUT_FILE}")
print(f"  Sheets: 测试用例 (60 rows), VTable列定义一览表, 汇总")
