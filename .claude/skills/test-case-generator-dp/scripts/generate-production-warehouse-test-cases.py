#!/usr/bin/env python3
# 生产缴库明细 - 测试用例导出
# 基于实际浏览器交互生成

# ============================================================
# 可配置变量
# ============================================================
ENTERPRISE_PREFIX = "NB"
DEFAULT_AUTHOR    = "Hooplus1ce"
MODULE_NAME       = "生产管理_生产缴库明细"
MODULE_LEVEL1     = "生产管理"
MODULE_LEVEL2     = "生产缴库明细"
MODULE_PINYIN    = "SCJKMX"
OUTPUT_DIR        = "."

# ============================================================
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ===================== Sheet 1: 测试用例 =====================
ws1 = wb.active
ws1.title = "测试用例"

HEADERS_19 = [
    "用例编号", "用例标题", "级别", "验证点", "一级模块", "二级模块",
    "测试类型", "功能", "前置条件", "测试步骤", "测试数据", "预期结果",
    "测试结果", "执行人", "执行时间", "编写人", "编写时间", "备注", "自动化建议"
]
ws1.append(HEADERS_19)

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 优先级配色
PRIORITY_STYLES = {
    "高级": (
        PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        Font(color="9C0006", name="微软雅黑")
    ),
    "中级": (
        PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        Font(color="9C5700", name="微软雅黑")
    ),
    "低级": (
        PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        Font(color="006100", name="微软雅黑")
    ),
}

def apply_priority_style(cell):
    style = PRIORITY_STYLES.get(cell.value)
    if style:
        cell.fill, cell.font = style

# ============================================================
# 测试用例数据 - 基于真实浏览器交互生成
# ============================================================
test_cases = [
    # ====================== F: 筛选查询类 ======================
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F001",
        "按缴入仓库筛选缴库记录",
        "低级",
        "筛选后表格仅显示指定仓库的数据",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "查询",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载完成",
        "1. 在缴入仓库字段下拉选择「样品仓」\n2. 点击查询按钮",
        "缴入仓库:样品仓",
        "表格仅显示缴入仓库为「样品仓」的记录，所有行缴入仓库列值均为「样品仓」，无不符合记录",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "使用 scan_filter_fields 获取字段矩阵，click 点击缴入库下拉选择选项，click_xy 点击查询按钮坐标，get_table_values 验证缴入仓库列值"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F002",
        "按制单日期范围筛选缴库记录",
        "低级",
        "筛选后表格仅显示指定日期范围内的数据",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "查询",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载完成",
        "1. 在制单时间开始日期选择「2026-07-01」\n2. 在制单时间结束日期选择「2026-07-05」\n3. 点击查询按钮",
        "制单时间开始:2026-07-01\n制单时间结束:2026-07-05",
        "表格仅显示制单时间在 2026-07-01 至 2026-07-05 范围内的记录，所有行制单时间均在此区间内",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "使用 select_date_range 设置制单时间范围，click_xy 点击查询按钮，get_table_values 扫描制单时间列验证"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F003",
        "重置按钮清空所有筛选条件",
        "低级",
        "点击重置后所有筛选字段恢复为空，表格显示全部数据",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "重置",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 缴入仓库已设置为「样品仓」，制单时间已设置日期范围",
        "1. 点击重置按钮",
        "",
        "所有筛选字段输入框清空，表格显示全部缴库记录，行数恢复为未筛选时数量",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "click_xy 点击重置按钮坐标，scan_filter_fields 验证所有字段 value 为空，scan_table 统计表格行数"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F004",
        "查询按钮空条件刷新表格",
        "低级",
        "不设置筛选条件点击查询，表格正常刷新无报错",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "查询",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 所有筛选字段为空",
        "1. 直接点击查询按钮",
        "",
        "表格正常刷新显示全部数据，无报错弹窗，无网络异常提示",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 network/notification/message，click_xy 点击查询，observe_wait 观察返回信号，listen_wait 验证接口 200 响应"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F005",
        "审批状态筛选功能验证",
        "低级",
        "按审批状态筛选显示对应状态的记录",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "查询",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 在审批状态字段下拉选择「已通过」\n2. 点击查询按钮",
        "审批状态:已通过",
        "表格仅显示审批状态为「已通过」的记录，所有行审批状态列均为「已通过」",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "click 点击审批状态下拉，click 选择下拉选项，click_xy 点击查询按钮，get_table_values 验证审批状态列"
    ],
    
    # ====================== I: 按钮交互类 ======================
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I001",
        "未勾选行点击打印按钮触发提示",
        "低级",
        "未勾选任何行时点击打印弹出提示",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "异常测试", "打印",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
        "1. 点击顶部工具栏「打印」按钮",
        "",
        "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 notification/message，click_xy 点击打印按钮坐标，observe_wait 验证提示文案匹配"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I002",
        "未勾选行点击批量打印触发提示",
        "低级",
        "未勾选任何行时点击批量打印弹出提示",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "异常测试", "批量打印",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
        "1. 点击顶部工具栏「批量打印」按钮",
        "",
        "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 notification/message，click_xy 点击批量打印按钮坐标，observe_wait 验证提示文案"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I003",
        "未勾选行点击添加生产缴库单触发提示",
        "低级",
        "未勾选任何行时点击添加生产缴库单弹出提示",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "异常测试", "新增",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
        "1. 点击顶部工具栏「添加生产缴库单」按钮",
        "",
        "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 notification/message，click_xy 点击添加生产缴库单按钮，observe_wait 验证提示文案"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I004",
        "未勾选行点击流程设置触发提示",
        "低级",
        "未勾选任何行时点击流程设置弹出提示",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "异常测试", "流程设置",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
        "1. 点击顶部工具栏「流程设置」按钮",
        "",
        "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 notification/message，click_xy 点击流程设置按钮，observe_wait 验证提示文案"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I005",
        "导出按钮空条件点击验证",
        "低级",
        "未设置筛选条件点击导出按钮",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "导出",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 所有筛选字段为空",
        "1. 点击顶部工具栏「导出」按钮",
        "",
        "触发导出接口请求，无报错提示，浏览器弹出文件下载提示",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 network/notification，listen_start 监听 gateway，click_xy 点击导出按钮，listen_wait 验证导出接口响应"
    ],
    
    # ====================== I: 表格操作类 ======================
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I006",
        "表格全选复选框功能验证",
        "低级",
        "点击表头复选框全选所有行，再次点击取消全选",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "表格交互",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 点击数据表格首列表头的复选框",
        "",
        "表格所有行记录被勾选，每行记录复选框均显示选中状态",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "click_table_cell 点击表头复选框单元格，scan_table 验证表格所有行 checkbox 状态"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I007",
        "表格列宽拖拽调整功能",
        "低级",
        "拖拽列边界可调整列宽，后续列自动移位",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "表格交互",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 鼠标移动到缴入仓库列的右边界\n2. 按住鼠标向右拖动扩大列宽\n3. 释放鼠标",
        "调整宽度:500px",
        "缴入仓库列宽从 120px 变为 500px，右侧所有列同步右移 380px，列内容显示完整无截断",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "resize_table_column 设置缴入仓库列为 500，scan_table 读取列坐标计算宽度变化量验证调整生效"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I008",
        "表格横向滚动查看右侧列",
        "低级",
        "横向滚动表格可查看视口外的列",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "表格交互",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 在数据表格区域向右横向滚动鼠标滚轮",
        "滚动距离:380px",
        "表格内容向左移动，原视口左侧的列移出视口，右侧的列进入视口可见区域",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "scan_table 获取初始列坐标，run_js 触发 wheel 事件向右滚动 380px，再次 scan_table 验证列坐标变化量匹配"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I009",
        "可跳转列单元格点击验证",
        "低级",
        "缴库单号等可点击列点击后可跳转或弹出详情",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "表格交互",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 点击第一行缴库单号列的单元格",
        "",
        "页面跳转到缴库单详情页或弹出缴库单详情弹窗，显示该条记录的完整信息",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "observe_start 监听 modal/tab/url 变化，click_table_cell 点击缴库单号单元格，observe_wait 验证跳转/弹窗"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I010",
        "表头筛选图标点击弹出筛选菜单",
        "低级",
        "点击列头筛选图标弹出对应列的筛选菜单",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "表格交互",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
        "1. 点击序号列头的筛选图标",
        "",
        "弹出该列的筛选下拉菜单，显示可选择的筛选条件选项",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "scan_table 获取筛选图标坐标，click_xy 点击筛选图标，run_js 检查 .vtable-filter-menu 元素存在"
    ],
    
    # ====================== P: 页面级测试 ======================
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P001",
        "页面首次加载数据表格渲染验证",
        "高级",
        "进入生产缴库明细页面后表格正常加载数据",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "页面加载",
        "1. 已登录 SCM 系统\n2. 在系统首页或其他模块页面",
        "1. 从左侧菜单点击「生产缴库明细」进入页面",
        "",
        "页面加载完成后数据表格显示缴库记录，表头列完整（26列有效数据列），无报错弹窗，无空白表格",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "enter_module 进入模块，observe_wait 等待加载，scan_table 验证列数量 >= 26，scan_page_elements 验证无报错弹窗"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P002",
        "筛选区内联模式显示验证",
        "中级",
        "筛选区以内联模式显示，所有筛选字段在页面内可见",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "页面布局",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面",
        "1. 观察筛选区显示模式\n2. 如为弹窗模式则切换为内联模式",
        "",
        "筛选区所有字段在页面内直接显示，点击筛选按钮不弹出高级搜索弹窗，显示「收起▲」按钮",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "dom_tree 检查 .page-query 结构，检查是否有 .legions-pro-quick-filter-remaining 元素存在，验证按钮文本为「收起▲」"
    ],
    [
        f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P003",
        "筛选字段矩阵完整性验证",
        "低级",
        "所有筛选字段正常显示，操作符下拉选项完整",
        MODULE_LEVEL1, MODULE_LEVEL2,
        "功能测试", "页面布局",
        "1. 已登录 SCM 系统\n2. 在生产缴库明细页面",
        "1. 扫描所有筛选字段\n2. 检查每个字段的操作符下拉选项",
        "",
        "共显示 25 个筛选字段，每个字段的操作符下拉选项完整（等于/包含/范围等操作符正确显示）",
        "", "", "",
        DEFAULT_AUTHOR, date.today().isoformat(), "",
        "scan_filter_fields 扫描所有筛选字段，验证字段数量 = 25，验证每个字段的 operatorOptions 非空"
    ]
]

# 写入用例数据
for row_data in test_cases:
    ws1.append(row_data)

# 应用单元格格式
for row_idx in range(2, ws1.max_row + 1):
    for col_idx in range(1, ws1.max_column + 1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx in (1, 2, 3, 5, 6, 7, 8, 13, 14, 15, 16, 17):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    apply_priority_style(ws1.cell(row=row_idx, column=3))

# 列宽设置
COL_WIDTHS = [18, 42, 12, 42, 18, 18, 18, 12, 50, 44, 44, 50, 10, 10, 12, 12, 12, 0, 58]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[chr(64 + i)].width = w

ws1.freeze_panes = "E2"
ws1.auto_filter.ref = f"A1:{ws1.cell(1, ws1.max_column).column_letter}{ws1.max_row}"

# ===================== Sheet 2: 测试数据 =====================
ws2 = wb.create_sheet(title="测试数据")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11, name="微软雅黑")
header2_font = Font(bold=True, size=10, name="微软雅黑")

def write_section(ws, start_row, title, headers, data_rows):
    col_count = len(headers)
    # 标题行
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = section_fill
    title_cell.font = section_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # 表头行
    hdr_row = start_row + 1
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = header2_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    # 数据行
    for row_idx, row_data in enumerate(data_rows, hdr_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return hdr_row + 1 + len(data_rows) + 2

row = 1
row = write_section(ws2, row, "3.1 测试数据配置",
    ["序号", "系统编号", "一级模块", "二级模块", "功能",
     "用例标题", "级别", "测试类型", "编写人", "编写日期"], [])

row = write_section(ws2, row, "3.2 筛选字段说明",
    ["字段名称", "输入方式", "对应测试数据", "说明"],
    [
        ["重量", "文本输入", "", "数值比较操作符下拉"],
        ["应生产数量", "文本输入", "", "数值比较操作符下拉"],
        ["批次号", "文本输入", "", "包含/等于等操作符"],
        ["缴入仓库", "下拉选择", "样品仓、成品仓、半成品仓、压铸仓库等", "16个仓库选项"],
        ["销售订单", "文本输入", "", "包含/等于等操作符"],
        ["客户名称", "文本输入", "", "包含/等于等操作符"],
        ["客户编码", "文本输入", "", "包含/等于等操作符"],
        ["缴库单号", "文本输入", "", "包含/等于等操作符"],
        ["缴库订单号", "文本输入", "", "包含/等于等操作符"],
        ["审批状态", "下拉选择", "审批中、已通过、已驳回、已删除", "4个状态选项"],
        ["审批人", "下拉选择", "", "用户列表选择"],
        ["商品识别码", "文本输入", "", "包含/等于等操作符"],
        ["商品名称", "文本输入", "", "包含/等于等操作符"],
        ["商品规格", "文本输入", "", "包含/等于等操作符"],
        ["加工方式", "下拉选择", "自制、外购、外发加工等", "6个加工方式选项"],
        ["中类名称", "文本输入", "", "包含/等于等操作符"],
        ["关联单号", "文本输入", "", "包含/等于等操作符"],
        ["生产部门", "文本输入", "", "包含/等于等操作符"],
        ["备注", "文本输入", "", "包含/等于等操作符"],
        ["制单时间", "日期范围", "2026-07-01 ~ 2026-07-05", "范围操作符"],
        ["缴库时间", "日期范围", "", "范围操作符"],
        ["缴库数量", "文本输入", "", "数值比较操作符"],
        ["经办人", "下拉选择", "", "用户列表选择"],
        ["入库状态", "下拉选择", "已缴库、待缴库", "2个状态选项"],
    ])

row = write_section(ws2, row, "3.3 各页签工具栏按钮一览表",
    ["按钮名称", "是否需要勾选行", "功能说明"],
    [
        ["导 出", "否", "导出当前筛选结果"],
        ["打 印", "是", "打印选中记录"],
        ["批量打印", "是", "批量打印选中记录"],
        ["添加生产缴库单", "是", "新增缴库单（需先选行）"],
        ["流程设置", "是", "设置审批流程（需先选行）"],
        ["查询", "否", "按筛选条件查询"],
        ["重置", "否", "清空所有筛选条件"],
    ])

row = write_section(ws2, row, "3.4 VTable列定义一览表",
    ["列标题", "列类型", "可交互", "说明"],
    [
        ["复选框", "checkbox", "是", "首列全选/行勾选"],
        ["序号", "text", "是", "序号显示，列头有筛选图标"],
        ["缴库单号", "link", "是", "可点击跳转详情"],
        ["缴库订单号", "link", "是", "可点击跳转详情"],
        ["审批状态", "link", "是", "可点击跳转详情"],
        ["审批人", "text", "是", "列头有筛选图标"],
        ["缴入仓库", "text", "是", "列头有筛选图标"],
        ["销售订单", "text", "是", "列头有筛选图标"],
        ["客户名称", "text", "是", "列头有筛选图标"],
        ["客户编码", "text", "是", "列头有筛选图标"],
        ["应生产数量", "text", "是", "列头有筛选图标"],
        ["批次号", "text", "是", "列头有筛选图标"],
        ["重量", "text", "是", "列头有筛选图标"],
        ["商品识别码", "link", "是", "可点击跳转详情"],
        ["商品规格", "text", "是", "列头有筛选图标"],
        ["商品名称", "text", "是", "列头有筛选图标"],
        ["加工方式", "link", "是", "可点击跳转详情"],
        ["中类名称", "text", "是", "列头有筛选图标"],
        ["关联单号", "link", "是", "可点击跳转详情"],
        ["生产部门", "text", "是", "列头有筛选图标"],
        ["备注", "text", "是", "列头有筛选图标"],
        ["制单时间", "link", "是", "可点击跳转详情"],
        ["缴库时间", "link", "是", "可点击跳转详情"],
        ["缴库数量", "text", "是", "列头有筛选图标"],
        ["经办人", "text", "是", "列头有筛选图标"],
        ["入库状态", "link", "是", "可点击跳转详情"],
    ])

# ===================== 保存 =====================
filename = f"测试用例_{MODULE_NAME}_{date.today().isoformat()}.xlsx"
os.makedirs(OUTPUT_DIR, exist_ok=True)
filepath = os.path.join(OUTPUT_DIR, filename)
try:
    wb.save(filepath)
    print(f"✅ 已生成: {os.path.abspath(filepath)}")
except PermissionError:
    fallback = os.path.join(".", filename)
    wb.save(fallback)
    print(f"⚠️ 降级保存至: {os.path.abspath(fallback)}")
