#!/usr/bin/env python3
"""
通用测试用例Excel生成器
功能：读取JSON格式的测试用例数据 → 生成标准化Excel文件
用法：uv run python .claude/skills/test-case-generator-dp/scripts/generate_from_json.py <json_file1> <json_file2> ...
"""

import os
import json
import glob
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_LEVEL1_PINYIN_MAP = {
    "生产管理": "SCGL",
    "协同管理": "XTGL",
    "基础数据": "JCSJ",
    "采购管理": "CGGL",
    "销售管理": "XSGL",
    "仓储管理": "CCGL",
    "系统管理": "XTGL",
}

# ============================================================
# 样式定义
header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11, name="微软雅黑")
header2_font = Font(bold=True, size=10, name="微软雅黑")

PRIORITY_STYLES = {
    "高级": (
        PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        Font(color="9C0006", name="微软雅黑")
    ),
    "中级": (
        PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        Font(color="9C5700", name="微软雅黑")
    ),
    "低级": (
        PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        Font(color="006100", name="微软雅黑")
    ),
}
# 企业标准测试用例Excel列定义（19列）
HEADERS_19 = [
    "用例编号", "用例标题", "级别", "一级模块", "二级模块", "测试类型",
    "功能", "验证点", "前置条件", "测试步骤", "测试数据", "预期结果",
    "测试结果", "执行人", "执行时间", "编写人", "编写时间", "备注", "自动化建议"
]

# 列宽设置（优化可读性，测试数据和预期结果加宽）
COL_WIDTHS = [18, 42, 12, 18, 18, 18, 12, 42, 50, 44, 44, 60, 10, 10, 12, 12, 12, 0, 58]

CASE_ID_RE = re.compile(r"^([A-Za-z]+)(\d+)$")
LEADING_NUMBER_RE = re.compile(r"^\s*\d+\s*[\.、)]\s*")


def natural_case_id_key(case_id):
    """用例编号自然排序：F2 在 F10 前。"""
    text = str(case_id or "")
    match = CASE_ID_RE.match(text)
    if match:
        return (match.group(1), int(match.group(2)), text)
    return (text, 10**9, text)


def sort_test_cases(test_cases):
    """按功能维度稳定排序，保证相同功能连续写入 Excel。"""
    function_order = {}
    for case in sorted(test_cases, key=lambda item: item.get("_source_order", 0)):
        function_name = case.get("function", "")
        if function_name not in function_order:
            function_order[function_name] = len(function_order)

    return sorted(
        test_cases,
        key=lambda case: (
            function_order.get(case.get("function", ""), len(function_order)),
            natural_case_id_key(case.get("case_id", "")),
            case.get("_source_order", 0),
        ),
    )


def strip_leading_number(text):
    """去掉源 JSON 中手写的步骤编号，避免导出时出现 1. 1. xxx。"""
    return LEADING_NUMBER_RE.sub("", str(text or "")).strip()


def numbered_lines(items):
    """按 Excel 标准输出连续编号列表。"""
    return "\n".join(
        f"{i + 1}. {strip_leading_number(item)}"
        for i, item in enumerate(items or [])
    )

def apply_priority_style(cell):
    """应用优先级颜色样式"""
    style = PRIORITY_STYLES.get(cell.value)
    if style:
        cell.fill, cell.font = style


def write_section(ws, start_row, title, headers, data_rows):
    """写入Sheet2的各个section"""
    col_count = max(len(headers), 6)
    # 标题行
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = section_fill
    title_cell.font = section_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    # 表头行
    hdr_row = start_row + 1
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = header2_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    # 数据行
    for row_idx, row_data in enumerate(data_rows, hdr_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return hdr_row + 1 + len(data_rows) + 2


def load_testcases_from_json(json_files):
    """从多个JSON文件加载测试用例"""
    all_cases = []
    module_info = None

    for fpath in sorted(json_files):
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 取第一个文件的module_info作为基准
        if module_info is None and "module_info" in data:
            module_info = data["module_info"]

        cases = data.get("test_cases", [])
        for case in cases:
            loaded_case = dict(case)
            loaded_case["_source_order"] = len(all_cases)
            all_cases.append(loaded_case)

        if cases:
            print(f"✅ 加载: {os.path.basename(fpath)} → {len(cases)} 条用例")

    return module_info, all_cases


def build_excel(module_info, test_cases, output_dir=None, custom_filename=None):
    """构建Excel文件"""
    test_cases = sort_test_cases(test_cases)

    # 从module_info提取配置
    prefix = module_info.get("enterprise_prefix", "NB")
    pinyin = module_info.get("module_pinyin", "TEST")
    level1 = module_info.get("module_level1", "")
    level2 = module_info.get("module_level2", "")
    author = module_info.get("author", "")
    write_date = date.today().isoformat()

    if output_dir is None or output_dir == ".":
        level1_pinyin = _LEVEL1_PINYIN_MAP.get(level1.strip()) if level1 else None
        if level1_pinyin:
            folder_name = f"{level1_pinyin}_{pinyin}"
        else:
            folder_name = pinyin
        output_dir = os.path.join("test_cases", folder_name)

    wb = Workbook()
    
    # ===================== Sheet 1: 测试用例 =====================
    ws1 = wb.active
    ws1.title = "测试用例"
    
    # 写入表头
    ws1.append(HEADERS_19)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 写入用例数据（列顺序与 HEADERS_19 完全一致）
    for case in test_cases:
        case_id = f"{prefix}_{pinyin}_{case['case_id']}"
        
        # 处理多行文本
        preconditions = numbered_lines(case.get("preconditions", []))
        test_steps = numbered_lines(case.get("test_steps", []))
        
        # 测试数据格式：匹配test.xlsx风格 - 首行固定"表单数据" + 中文冒号键值对
        test_data_dict = case.get("test_data", {})
        if test_data_dict:
            test_data_parts = ["表单数据"]
            for k, v in test_data_dict.items():
                test_data_parts.append(f"{k}：{v}")  # 注意：中文冒号！
            test_data = "\n".join(test_data_parts)
        else:
            test_data = ""
        
        # 预期结果已经是编号列表格式，直接使用
        expected_result = case["expected_result"]
        
        row_data = [
            case_id,                    # 1. 用例编号
            case["case_title"],         # 2. 用例标题
            case["priority"],           # 3. 级别
            level1,                     # 4. 一级模块
            level2,                     # 5. 二级模块
            case["test_type"],          # 6. 测试类型
            case["function"],           # 7. 功能
            case["verify_point"],       # 8. 验证点
            preconditions,              # 9. 前置条件
            test_steps,                 # 10. 测试步骤
            test_data,                  # 11. 测试数据（表单数据格式）
            expected_result,            # 12. 预期结果（编号列表格式）
            "", "", "",                 # 13.测试结果 /14.执行人 /15.执行时间
            author,                     # 16. 编写人
            write_date,                 # 17. 编写时间
            "",                         # 18. 备注
            case.get("automation_suggestion", "")  # 19. 自动化建议
        ]
        ws1.append(row_data)
    
    # 应用单元格格式
    for row_idx in range(2, ws1.max_row + 1):
        for col_idx in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            # 居中对齐的列（第3列是级别）
            if col_idx in (1, 3, 4, 5, 6, 7, 13, 14, 15, 16, 17):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        apply_priority_style(ws1.cell(row=row_idx, column=3))  # 第3列是级别
    
    # 列宽设置
    for i, w in enumerate(COL_WIDTHS, 1):
        ws1.column_dimensions[chr(64 + i)].width = w
    
    ws1.auto_filter.ref = f"A1:{ws1.cell(1, ws1.max_column).column_letter}{ws1.max_row}"
    
    # ===================== Sheet 2: 测试数据 =====================
    ws2 = wb.create_sheet(title="测试数据")
    row = 1
    
    # 统计汇总
    total = len(test_cases)
    priority_count = {"高级": 0, "中级": 0, "低级": 0}
    type_count = {}
    func_count = {}
    
    for case in test_cases:
        priority_count[case["priority"]] = priority_count.get(case["priority"], 0) + 1
        type_count[case["test_type"]] = type_count.get(case["test_type"], 0) + 1
        func_count[case["function"]] = func_count.get(case["function"], 0) + 1
    
    row = write_section(ws2, row, "1. 测试用例统计汇总",
        ["统计项", "数值", "占比"], [
            ["总用例数", str(total), "100%"],
            ["高级用例", str(priority_count["高级"]), f"{priority_count['高级']/total*100:.1f}%"],
            ["中级用例", str(priority_count["中级"]), f"{priority_count['中级']/total*100:.1f}%"],
            ["低级用例", str(priority_count["低级"]), f"{priority_count['低级']/total*100:.1f}%"],
            ["", "", ""],
        ] + [[k, str(v), f"{v/total*100:.1f}%"] for k, v in type_count.items()])
    
    row = write_section(ws2, row, "2. 测试数据配置",
        ["序号", "系统编号", "一级模块", "二级模块", "功能",
         "用例标题", "级别", "测试类型", "编写人", "编写日期"], [])
    
    # 3. 筛选字段说明（如果有）
    filter_fields = []
    if "meta_data" in test_cases[0] if test_cases else False:
        filter_fields = test_cases[0]["meta_data"].get("filter_fields", [])
    
    row = write_section(ws2, row, "3. 筛选字段说明",
        ["字段名称", "输入方式", "对应测试数据", "说明"], filter_fields or [
            ["（探索过程中自动填充）", "", "", ""],
        ])
    
    row = write_section(ws2, row, "4. 工具栏按钮一览表",
        ["按钮名称", "是否需要勾选行", "功能说明"], [
            ["（探索过程中自动填充）", "", ""],
        ])
    
    row = write_section(ws2, row, "5. VTable列定义一览表",
        ["列标题", "列类型", "可交互", "说明"], [
            ["（探索过程中自动填充）", "", "", ""],
        ])
    
    # ===================== 保存 =====================
    module_name = module_info.get("module_name", "测试用例")
    if custom_filename:
        filename = custom_filename
    else:
        filename = f"测试用例_{module_name}_{date.today().isoformat()}.xlsx"
    
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    
    try:
        wb.save(filepath)
        print(f"\n{'='*60}")
        print(f"✅ 测试用例生成完成！")
        print(f"📁 文件路径: {os.path.abspath(filepath)}")
        print(f"📊 用例数量: {len(test_cases)} 条")
        print(f"  - 高级: {priority_count['高级']} 条")
        print(f"  - 中级: {priority_count['中级']} 条")
        print(f"  - 低级: {priority_count['低级']} 条")
        print(f"{'='*60}")
        return filepath
    except PermissionError:
        stem, ext = os.path.splitext(filename)
        fallback_name = f"{stem}_修正版{ext}"
        fallback = os.path.join(output_dir, fallback_name)
        index = 2
        while os.path.exists(fallback):
            fallback_name = f"{stem}_修正版{index}{ext}"
            fallback = os.path.join(output_dir, fallback_name)
            index += 1
        wb.save(fallback)
        print(f"⚠️ 降级保存至: {os.path.abspath(fallback)}")
        return fallback


def main():
    import sys
    
    if len(sys.argv) < 2:
        print("用法: uv run python .claude/skills/test-case-generator-dp/scripts/generate_from_json.py <json_file1> [<json_file2> ...]")
        print("\n示例:")
        print("  # 单个文件")
        print("  uv run python .claude/skills/test-case-generator-dp/scripts/generate_from_json.py test_cases/SCJKMX/SCJKMX_01_筛选查询类.json")
        print("  # 多个文件合并")
        print("  uv run python .claude/skills/test-case-generator-dp/scripts/generate_from_json.py test_cases/SCJKMX/SCJKMX_*.json")
        print("  # 其他模块")
        print("  uv run python .claude/skills/test-case-generator-dp/scripts/generate_from_json.py test_cases/OTHER/OTHER_*.json")
        return
    
    # 处理通配符
    json_files = []
    for arg in sys.argv[1:]:
        if "*" in arg or "?" in arg:
            json_files.extend(glob.glob(arg))
        else:
            json_files.append(arg)
    
    if not json_files:
        print("❌ 未找到任何JSON文件")
        return
    
    module_info, test_cases = load_testcases_from_json(json_files)
    
    if not test_cases:
        print("❌ 未加载到任何测试用例")
        return
    
    build_excel(module_info, test_cases)


if __name__ == "__main__":
    main()
