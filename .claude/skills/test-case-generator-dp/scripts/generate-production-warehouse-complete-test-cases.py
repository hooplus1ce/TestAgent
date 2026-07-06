#!/usr/bin/env python3
# 生产缴库明细 - 完整测试用例导出
# 基于企业级系统测试规范，覆盖所有测试类型

# ============================================================
# 可配置变量
# ============================================================
ENTERPRISE_PREFIX = "NB"
DEFAULT_AUTHOR    = "Hooplus1ce"
MODULE_NAME       = "生产管理_生产缴库明细"
MODULE_LEVEL1     = "生产管理"
MODULE_LEVEL2     = "生产缴库明细"
MODULE_PINYIN    = "SCJKMX"
OUTPUT_DIR        = "."

# ============================================================
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ===================== Sheet 1: 测试用例 =====================
ws1 = wb.active
ws1.title = "测试用例"

HEADERS_19 = [
    "用例编号", "用例标题", "级别", "验证点", "一级模块", "二级模块",
    "测试类型", "功能", "前置条件", "测试步骤", "测试数据", "预期结果",
    "测试结果", "执行人", "执行时间", "编写人", "编写时间", "备注", "自动化建议"
]
ws1.append(HEADERS_19)

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 优先级配色
PRIORITY_STYLES = {
    "高级": (
        PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        Font(color="9C0006", name="微软雅黑")
    ),
    "中级": (
        PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        Font(color="9C5700", name="微软雅黑")
    ),
    "低级": (
        PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        Font(color="006100", name="微软雅黑")
    ),
}

def apply_priority_style(cell):
    style = PRIORITY_STYLES.get(cell.value)
    if style:
        cell.fill, cell.font = style

# ============================================================
# 完整测试用例 - 覆盖所有测试类型
# ============================================================
test_cases = []

# ========== F001-F025: 筛选查询功能测试 ==========
# F001-F009: 单字段功能测试
test_cases.extend([
    # F001
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F001", "按缴入仓库精准筛选缴库记录", "低级", "筛选后表格仅显示指定仓库的数据", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载完成",
     "1. 在缴入仓库字段下拉选择「样品仓」\n2. 点击查询按钮",
     "缴入仓库:样品仓",
     "表格仅显示缴入仓库为「样品仓」的记录，所有行缴入仓库列值均为「样品仓」，无不符合记录，筛选后记录数大于等于0",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "scan_filter_fields 获取字段矩阵，click 选择下拉选项，click_xy 点击查询，get_table_values 验证缴入仓库列值，统计行数变化"],
    # F002
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F002", "按制单日期范围筛选缴库记录", "低级", "筛选后表格仅显示指定日期范围内的数据", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载完成",
     "1. 在制单时间开始日期选择「2026-07-01」\n2. 在制单时间结束日期选择「2026-07-05」\n3. 点击查询按钮",
     "制单时间开始:2026-07-01\n制单时间结束:2026-07-05",
     "表格仅显示制单时间在 2026-07-01 至 2026-07-05 范围内的记录，所有行制单时间均在此区间内，无超出范围记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "select_date_range 设置范围，click_xy 点击查询，get_table_values 扫描制单时间列，逐行验证时间区间"],
    # F003
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F003", "重置按钮清空所有筛选条件", "低级", "点击重置后所有筛选字段恢复为空，表格显示全部数据", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "重置",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 缴入仓库已设置为「样品仓」，制单时间已设置日期范围",
     "1. 点击重置按钮",
     "",
     "所有筛选字段输入框清空，缴入仓库恢复为「请选择」，日期范围清空，表格显示全部缴库记录，行数恢复为未筛选时数量",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击重置按钮，scan_filter_fields 验证所有字段 value 为空，scan_table 统计表格行数并与基准行数比较"],
    # F004
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F004", "查询按钮空条件刷新表格", "低级", "不设置筛选条件点击查询，表格正常刷新无报错", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 所有筛选字段为空",
     "1. 直接点击查询按钮",
     "",
     "表格正常刷新显示全部数据，无报错弹窗，无网络异常提示，接口响应状态码为200",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 network/notification，click_xy 点击查询，observe_wait 观察信号，listen_wait 验证接口200响应"],
    # F005
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F005", "审批状态筛选功能验证", "低级", "按审批状态筛选显示对应状态的记录", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 在审批状态字段下拉选择「已通过」\n2. 点击查询按钮",
     "审批状态:已通过",
     "表格仅显示审批状态为「已通过」的记录，所有行审批状态列均为「已通过」，无其他状态记录混入",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 点击审批状态下拉，click 选择选项，click_xy 点击查询，get_table_values 验证审批状态列"],
    # F006
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F006", "按重量小于指定值筛选", "低级", "数值比较操作符「小于」功能正常", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 重量字段操作符选择「小于」\n2. 输入框输入「100」\n3. 点击查询按钮",
     "重量 < 100",
     "表格仅显示重量小于100的记录，所有行重量列数值均 < 100，无大于等于100的记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「小于」操作符，input 输入100，click_xy 点击查询，get_table_values 获取重量列，逐行验证数值比较"],
    # F007
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F007", "按缴库数量大于指定值筛选", "低级", "数值比较操作符「大于」功能正常", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴库数量字段操作符选择「大于」\n2. 输入框输入「50」\n3. 点击查询按钮",
     "缴库数量 > 50",
     "表格仅显示缴库数量大于50的记录，所有行缴库数量列数值均 > 50，无小于等于50的记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「大于」操作符，input 输入50，click_xy 点击查询，get_table_values 获取缴库数量列，逐行验证"],
    # F008
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F008", "商品名称模糊包含筛选", "低级", "文本模糊匹配「包含」操作符功能正常", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 商品名称字段操作符选择「包含」\n2. 输入框输入「轴承」\n3. 点击查询按钮",
     "商品名称包含:轴承",
     "表格仅显示商品名称包含「轴承」关键字的记录，所有行商品名称均包含「轴承」字符串，无不匹配记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「包含」操作符，input 输入「轴承」，click_xy 点击查询，get_table_values 获取商品名称列，逐行验证字符串包含"],
    # F009
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F009", "加工方式枚举筛选", "低级", "枚举下拉筛选功能正常，所有选项可选择", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 加工方式字段下拉选择「自制」\n2. 点击查询按钮",
     "加工方式:自制",
     "表格仅显示加工方式为「自制」的记录，下拉选项列表完整包含「自制/外购/外发加工」等全部6个选项",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 打开加工方式下拉，verify 下拉选项数量=6，click 选择「自制」，click_xy 点击查询，get_table_values 验证加工方式列"],
])

# F010-F018: 边界值测试
test_cases.extend([
    # F010
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F010", "重量字段输入0边界值筛选", "低级", "数值输入边界值0筛选功能正常", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 重量字段操作符选择「等于」\n2. 输入框输入「0」\n3. 点击查询按钮",
     "重量 = 0",
     "系统正常处理0值输入，无报错弹窗，表格显示重量等于0的记录或显示空结果（无匹配时）",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「等于」操作符，input 输入0，click_xy 点击查询，observe_wait 验证无报错，scan_table 验证表格状态正常"],
    # F011
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F011", "重量字段输入极大值边界测试", "低级", "数值输入极大值系统健壮性验证", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 重量字段操作符选择「大于」\n2. 输入框输入「999999999999」（12位最大值）\n3. 点击查询按钮",
     "重量 > 999999999999",
     "系统正常处理极大值输入，无溢出报错，无前端崩溃，表格显示空结果（无匹配时），接口响应正常",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「大于」操作符，input 输入12位最大值，click_xy 点击查询，observe_wait 验证无报错，listen_wait 验证接口响应正常"],
    # F012
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F012", "文本字段输入超长字符串边界测试", "低级", "文本输入超长字符串系统健壮性验证", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 商品名称字段操作符选择「等于」\n2. 输入框输入200个字符的超长字符串\n3. 点击查询按钮",
     "商品名称 = 200字符超长字符串",
     "系统正常处理超长输入，无截断报错，无前端崩溃，表格显示空结果（无匹配时），接口响应状态正常",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「等于」操作符，input 输入200字符字符串，click_xy 点击查询，observe_wait 验证无报错，scan_table 验证表格正常渲染"],
    # F013
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F013", "日期范围同一天边界测试", "低级", "日期范围开始日期=结束日期边界情况", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 制单时间开始日期选择「2026-07-03」\n2. 制单时间结束日期选择「2026-07-03」\n3. 点击查询按钮",
     "制单时间:2026-07-03 至 2026-07-03",
     "系统正常处理同一天日期范围，表格显示当天的缴库记录，无报错，日期筛选逻辑正确",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "select_date_range 设置同一天，click_xy 点击查询，get_table_values 验证制单时间均为2026-07-03"],
    # F014
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F014", "日期范围开始大于结束异常测试", "低级", "日期范围非法输入系统处理验证", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 制单时间开始日期选择「2026-07-10」\n2. 制单时间结束日期选择「2026-07-01」\n3. 点击查询按钮",
     "制单时间开始>结束",
     "系统正确识别非法日期范围，弹出友好提示「开始日期不能大于结束日期」，不执行查询，表格保持原数据",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "select_date_range 设置开始>结束，observe_start 监听 notification，click_xy 点击查询，observe_wait 验证提示文案，scan_table 验证表格数据未变"],
    # F015
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F015", "数值字段输入负数异常测试", "低级", "数值输入负数系统处理验证", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴库数量字段操作符选择「小于」\n2. 输入框输入「-100」\n3. 点击查询按钮",
     "缴库数量 < -100",
     "系统正常处理负数输入（业务无负数则显示空结果），无报错弹窗，无前端崩溃，接口响应正常",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「小于」操作符，input 输入-100，click_xy 点击查询，observe_wait 验证无报错，scan_table 验证表格状态正常"],
    # F016
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F016", "文本字段输入特殊字符异常测试", "低级", "SQL注入/XSS特殊字符系统处理验证", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 商品名称字段操作符选择「包含」\n2. 输入框输入「' OR 1=1 -- <script>alert(1)</script>」\n3. 点击查询按钮",
     "商品名称包含:SQL注入/XSS字符",
     "系统正确转义特殊字符，无SQL注入漏洞，无XSS弹窗，查询正常执行，无报错，系统安全机制生效",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「包含」操作符，input 输入特殊字符，click_xy 点击查询，observe_wait 验证无异常弹窗/报错，listen_wait 验证接口响应正常"],
    # F017
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F017", "无匹配结果空状态显示测试", "低级", "筛选无结果时空状态提示验证", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴库单号字段操作符选择「等于」\n2. 输入框输入「不存在的单号XXX123」\n3. 点击查询按钮",
     "缴库单号 = 不存在的单号XXX123",
     "表格显示空状态，友好提示「暂无数据」，分页显示0条记录，表格布局正常无错乱",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择「等于」操作符，input 输入不存在单号，click_xy 点击查询，scan_table 验证空状态，dom_tree 检查「暂无数据」提示存在"],
    # F018
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F018", "全选下拉所有仓库筛选测试", "低级", "下拉框全选选项功能验证", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴入仓库字段下拉勾选「全选」选项\n2. 点击查询按钮",
     "缴入仓库:全选",
     "表格显示全部仓库的缴库记录，与未筛选状态一致，全选选项功能正常，勾选状态正确显示",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 打开缴入仓库下拉，click 勾选「全选」，click_xy 点击查询，scan_table 验证行数与未筛选时一致"],
])

# F019-F025: 组合筛选与一致性测试
test_cases.extend([
    # F019
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F019", "仓库+审批状态双条件AND组合筛选", "中级", "多字段AND逻辑组合筛选功能正确", MODULE_LEVEL1, MODULE_LEVEL2, "组合条件测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴入仓库选择「样品仓」\n2. 审批状态选择「已通过」\n3. 点击查询按钮",
     "缴入仓库=样品仓 AND 审批状态=已通过",
     "表格仅显示同时满足两个条件的记录：缴入仓库=样品仓 且 审批状态=已通过，无单条件满足但另一条件不满足的记录",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择样品仓，click 选择已通过，click_xy 点击查询，get_table_values 逐行验证两个字段同时满足条件"],
    # F020
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F020", "仓库+日期范围三条件组合筛选", "中级", "三个字段AND逻辑组合筛选功能正确", MODULE_LEVEL1, MODULE_LEVEL2, "组合条件测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 缴入仓库选择「成品仓」\n2. 制单时间设置 2026-07-01 至 2026-07-15\n3. 审批状态选择「审批中」\n4. 点击查询按钮",
     "成品仓 AND 7月1-15日 AND 审批中",
     "表格仅显示同时满足三个条件的记录，三个筛选条件逻辑关系为AND，结果集是三个条件的交集",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 选择成品仓，select_date_range 设置日期，click 选择审批中，click_xy 点击查询，逐行验证三个字段同时满足"],
    # F021
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F021", "筛选结果排序一致性验证", "中级", "筛选后点击列排序，筛选条件保持有效", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "查询+排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 已按缴入仓库=样品仓筛选完成",
     "1. 点击制单时间列表头进行降序排序",
     "缴入仓库=样品仓",
     "排序后筛选条件仍然有效，所有记录仍然满足缴入仓库=样品仓，仅按制单时间降序重新排列，无其他仓库记录混入",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击制单时间列表头排序，get_table_values 验证缴入仓库列仍全部为样品仓，验证制单时间降序排列"],
    # F022
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F022", "筛选结果分页一致性验证", "中级", "筛选后翻页，筛选条件保持有效", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "查询+分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 已按审批状态=已通过筛选完成，结果多于1页",
     "1. 点击分页控件「下一页」按钮",
     "审批状态=已通过",
     "翻页后筛选条件仍然有效，第二页所有记录审批状态仍为「已通过」，无其他状态记录混入，分页计算正确",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击下一页按钮，observe_wait 等待加载，get_table_values 验证审批状态列仍全部为已通过"],
    # F023
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F023", "连续多次查询幂等性测试", "低级", "连续点击查询按钮结果一致，无重复数据", MODULE_LEVEL1, MODULE_LEVEL2, "幂等性测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 缴入仓库已设置为「成品仓」",
     "1. 快速连续点击查询按钮3次",
     "缴入仓库=成品仓",
     "三次查询结果完全一致，记录数相同，无重复数据，无数据丢失，接口去重/防抖机制生效，无并发异常",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 network，连续三次 click_xy 点击查询，listen_wait 捕获所有请求，比较三次响应数据完全一致"],
    # F024
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F024", "查询中点击重置状态回滚测试", "低级", "查询请求未完成时点击重置，状态正确回滚", MODULE_LEVEL1, MODULE_LEVEL2, "状态一致性测试", "重置",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 网络设置为慢速模拟",
     "1. 点击查询按钮\n2. 在请求未返回前立即点击重置按钮",
     "",
     "重置操作生效，所有筛选字段清空，表格最终显示全部数据，无半成功状态，无前端状态不一致",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 模拟网络延迟，click_xy 点击查询后立即 click_xy 点击重置，observe_wait 验证最终状态：筛选字段为空+表格显示全部数据"],
    # F025
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_F025", "筛选条件切换查询结果正确性", "低级", "多次切换不同筛选条件，结果正确无缓存污染", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "查询",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 筛选缴入仓库=样品仓，点击查询\n2. 记录结果数量\n3. 改为筛选缴入仓库=成品仓，点击查询\n4. 记录结果数量\n5. 再次改为筛选缴入仓库=样品仓，点击查询",
     "样品仓 → 成品仓 → 样品仓",
     "第三次查询结果与第一次完全一致，记录数相同，无第二次查询结果的缓存残留，条件切换干净",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "三次设置不同条件查询，scan_table 分别记录三次行数，验证第1次行数 = 第3次行数"],
])

# ========== I001-I040: 按钮交互与表格操作测试 ==========
# I001-I015: 按钮交互测试
test_cases.extend([
    # I001
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I001", "未勾选行点击打印按钮触发提示", "低级", "未勾选任何行时点击打印弹出「请选择一条数据」提示", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "打印",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
     "1. 点击顶部工具栏「打印」按钮",
     "",
     "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转，提示3秒后自动消失",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 notification/message，click_xy 点击打印按钮坐标，observe_wait 验证提示文案匹配，计时验证3秒后消失"],
    # I002
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I002", "未勾选行点击批量打印触发提示", "低级", "未勾选任何行时点击批量打印弹出提示", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "批量打印",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
     "1. 点击顶部工具栏「批量打印」按钮",
     "",
     "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 notification，click_xy 点击批量打印按钮，observe_wait 验证提示文案"],
    # I003
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I003", "未勾选行点击添加生产缴库单触发提示", "低级", "未勾选任何行时点击添加生产缴库单弹出提示", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "新增",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
     "1. 点击顶部工具栏「添加生产缴库单」按钮",
     "",
     "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 notification，click_xy 点击添加生产缴库单按钮，observe_wait 验证提示文案"],
    # I004
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I004", "未勾选行点击流程设置触发提示", "低级", "未勾选任何行时点击流程设置弹出提示", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "流程设置",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 未勾选任何行记录",
     "1. 点击顶部工具栏「流程设置」按钮",
     "",
     "页面弹出「请选择一条数据」的提示通知，无其他弹窗或跳转",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 notification，click_xy 点击流程设置按钮，observe_wait 验证提示文案"],
    # I005
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I005", "导出按钮空条件点击验证", "低级", "未设置筛选条件点击导出按钮", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "导出",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 所有筛选字段为空",
     "1. 点击顶部工具栏「导出」按钮",
     "",
     "触发导出接口请求，无报错提示，浏览器弹出文件下载提示，导出文件格式为Excel，包含所有字段",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 network，listen_start 监听 gateway，click_xy 点击导出按钮，listen_wait 验证导出接口响应200"],
    # I006
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I006", "勾选单行后点击打印功能验证", "低级", "勾选单行记录后点击打印正常跳转", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "打印",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 已勾选第1行记录",
     "1. 点击顶部工具栏「打印」按钮",
     "",
     "正常跳转到打印预览页面，打印内容与勾选记录一致，无报错提示",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 勾选第1行复选框，observe_start 监听 tab/url 变化，click_xy 点击打印，observe_wait 验证打印页面打开"],
    # I007
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I007", "勾选多行后点击批量打印功能验证", "低级", "勾选多行记录后点击批量打印正常处理", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "批量打印",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 已勾选第1、2、3行记录",
     "1. 点击顶部工具栏「批量打印」按钮",
     "勾选3条记录",
     "正常跳转到批量打印预览页面，打印内容包含所有勾选的3条记录，无遗漏无多余",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 分别勾选3行，observe_start 监听 tab/url 变化，click_xy 点击批量打印，observe_wait 验证打印页面包含3条记录"],
    # I008
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I008", "勾选1条后点击添加生产缴库单功能验证", "低级", "勾选单条记录后点击添加缴库单正常弹出详情页", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "新增",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 已勾选第1行记录",
     "1. 点击顶部工具栏「添加生产缴库单」按钮",
     "",
     "弹出添加生产缴库单弹窗/新标签页，自动带入勾选记录的相关数据，表单字段正确填充",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 勾选第1行，observe_start 监听 modal/tab 变化，click_xy 点击添加按钮，observe_wait 验证弹窗/新页面打开"],
    # I009
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I009", "勾选1条后点击流程设置功能验证", "低级", "勾选单条记录后点击流程设置正常弹出配置页", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "流程设置",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 已勾选第1行记录",
     "1. 点击顶部工具栏「流程设置」按钮",
     "",
     "弹出流程设置弹窗，显示该记录当前审批流程配置，可编辑流程节点和审批人",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 勾选第1行，observe_start 监听 modal 变化，click_xy 点击流程设置，observe_wait 验证流程设置弹窗打开"],
    # I010
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I010", "筛选后导出数据一致性验证", "低级", "筛选后导出的数据与页面显示数据一致", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "导出",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 已按缴入仓库=样品仓筛选完成，显示N条记录",
     "1. 点击顶部工具栏「导出」按钮\n2. 下载导出的Excel文件",
     "缴入仓库=样品仓",
     "导出的Excel文件包含N条记录，所有记录缴入仓库均为样品仓，与页面显示完全一致，字段列完整无缺失",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "按条件筛选后 scan_table 记录页面行数，click_xy 点击导出，download_by_browser 下载文件，读取Excel验证行数与页面一致，验证缴库仓库列均为样品仓"],
    # I011
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I011", "导出大数量记录性能测试", "中级", "导出超过1000条记录时系统性能与稳定性", MODULE_LEVEL1, MODULE_LEVEL2, "性能测试", "导出",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格有超过1000条记录",
     "1. 不设置筛选条件（导出全部）\n2. 点击顶部工具栏「导出」按钮",
     "1000+条记录",
     "系统正常处理大数量导出，无超时无崩溃，导出进度提示正常，最终文件完整下载，导出耗时 < 30秒",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 network，click_xy 点击导出，计时开始，listen_wait 等待接口响应，验证响应时间 < 30秒，验证文件下载完成"],
    # I012
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I012", "连续点击导出防抖测试", "低级", "快速连续点击导出按钮，仅触发一次导出", MODULE_LEVEL1, MODULE_LEVEL2, "幂等性测试", "导出",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 快速连续点击导出按钮3次",
     "",
     "防抖机制生效，仅触发1次导出请求，浏览器仅弹出1个下载提示，无重复导出文件，按钮点击期间禁用状态",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 network，快速连续3次 click_xy 点击导出，listen_wait 统计请求数量 = 1，验证按钮禁用期间点击无效"],
    # I013
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I013", "导出中关闭页面拦截测试", "低级", "导出请求未完成时关闭页面，弹出确认提示", MODULE_LEVEL1, MODULE_LEVEL2, "状态一致性测试", "导出",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 网络设置为慢速模拟",
     "1. 点击导出按钮\n2. 在请求未返回前尝试关闭浏览器标签页",
     "",
     "浏览器弹出「下载进行中，确定要离开吗？」确认提示，用户选择取消则继续导出，选择离开则中止导出",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 模拟网络延迟，click_xy 点击导出后立即 run_js 触发页面关闭，验证浏览器弹出离开确认提示"],
    # I014
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I014", "勾选全选后打印全部测试", "中级", "点击全选后打印，包含当前页所有记录", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "打印",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 当前页显示20条记录",
     "1. 点击表头全选复选框\n2. 点击打印按钮",
     "全选20条",
     "打印预览包含当前页全部20条记录，无遗漏，打印内容按表格顺序排列",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 点击表头全选，scan_table 验证所有行已勾选，click_xy 点击打印，验证打印预览记录数 = 20"],
    # I015
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I015", "取消勾选后按钮状态更新测试", "低级", "取消所有勾选后，需勾选按钮恢复未选中提示逻辑", MODULE_LEVEL1, MODULE_LEVEL2, "状态一致性测试", "交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 已勾选第1行记录",
     "1. 取消勾选第1行记录\n2. 点击打印按钮",
     "",
     "取消勾选后立即弹出「请选择一条数据」提示，与从未勾选状态一致，按钮逻辑正确识别勾选状态变化",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 勾选第1行，再次 click_table_cell 取消勾选，click_xy 点击打印，observe_wait 验证提示正常弹出"],
])

# I016-I040: 表格操作测试
test_cases.extend([
    # I016
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I016", "表格全选复选框功能验证", "低级", "点击表头复选框全选所有行，再次点击取消全选", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载，当前页20条记录",
     "1. 点击数据表格首列表头的复选框\n2. 验证所有行被勾选\n3. 再次点击表头复选框",
     "",
     "第一次点击后当前页所有20行记录被勾选，每行复选框显示选中状态；第二次点击后所有行取消勾选，全选状态正确切换",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 点击表头复选框，scan_table 验证所有行 checkbox 状态为选中，再次点击后验证全部取消选中"],
    # I017
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I017", "单行复选框独立状态验证", "低级", "单行勾选/取消不影响其他行，全选状态正确联动", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载\n4. 全选所有20行已勾选",
     "1. 取消勾选第5行记录\n2. 观察表头全选复选框状态",
     "",
     "取消第5行后，其他19行仍保持勾选状态，表头全选复选框变为「半选」状态（indeterminate），正确反映部分选中状态",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_table_cell 全选所有行，click_table_cell 取消第5行，run_js 检查表头复选框 indeterminate 属性为 true，scan_table 验证其他19行仍选中"],
    # I018
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I018", "表格列宽拖拽调整功能", "低级", "拖拽列边界可调整列宽，后续列自动移位", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 鼠标移动到缴入仓库列的右边界\n2. 按住鼠标向右拖动扩大列宽\n3. 释放鼠标",
     "调整宽度:500px",
     "缴入仓库列宽从120px变为500px，右侧所有列同步右移380px，列内容显示完整无截断，列边界对齐整齐",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "resize_table_column 设置缴入仓库列为500，scan_table 读取列坐标计算宽度变化量，验证调整量=380px，验证右侧列均同步移动"],
    # I019
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I019", "列宽调至最小边界测试", "低级", "列宽拖拽至最小允许宽度，列内容自适应换行或缩略", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 鼠标移动到缴入仓库列的右边界\n2. 按住鼠标向左拖动至最窄\n3. 释放鼠标",
     "最小列宽",
     "列宽达到系统预设最小值（如50px）后无法继续缩小，列内容自动换行或显示省略号，表格布局无错乱",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "resize_table_column 设置缴入仓库列为最小（如10px期望被限制为50px），scan_table 验证实际列宽=最小限制值，验证列内容显示正常无溢出"],
    # I020
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I020", "多列依次调整宽度稳定性测试", "中级", "连续调整多列宽度，所有调整均生效且无相互影响", MODULE_LEVEL1, MODULE_LEVEL2, "稳定性测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 调整缴入仓库列宽至400px\n2. 调整销售订单列宽至300px\n3. 调整客户名称列宽至350px\n4. 刷新页面",
     "三列依次调整",
     "三列宽度调整均正确生效，刷新后列宽保持调整后的值（如有本地存储记忆功能），表格布局整齐无重叠",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "依次 resize_table_column 三列，scan_table 分别验证每列宽度，run_js 刷新页面后再次验证列宽保持（如设计有记忆功能）"],
    # I021
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I021", "表格横向滚动查看右侧列", "低级", "横向滚动表格可查看视口外的列，滚动平滑", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载，总列宽超出视口",
     "1. 在数据表格区域向右横向滚动鼠标滚轮至最右端",
     "滚动距离:完整宽度",
     "表格内容向左移动，原视口左侧的列移出视口，最右侧列表头进入视口可见，滚动过程平滑无卡顿，列对齐正常",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "scan_table 获取初始列坐标，run_js 触发横向滚动至最右端，再次 scan_table 验证最右列（入库状态）x坐标进入视口范围"],
    # I022
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I022", "固定列横向滚动锁定测试", "低级", "横向滚动时首列复选框固定不移动", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载，首列设置为固定列",
     "1. 在数据表格区域向右横向滚动\n2. 观察首列位置",
     "横向滚动",
     "横向滚动过程中，首列（复选框列）始终固定在视口最左侧，不随滚动移动，其他列正常滚动，固定列边框清晰无重影",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 触发横向滚动，scan_table 验证首列x坐标保持不变，验证其他列x坐标均减小（向左移动）"],
    # I023
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I023", "横向滚动边界回弹测试", "低级", "滚动到左右边界时继续滚动无异常", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 横向滚动到最左端后继续向左滚动\n2. 横向滚动到最右端后继续向右滚动",
     "",
     "滚动到边界后继续滚动无JS报错，表格布局保持正常，无白屏或渲染错乱，边界回弹效果正常（如有）",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 滚动到最左后继续触发向左滚动事件，run_js 滚动到最右后继续触发向右滚动，observe_wait 验证无控制台报错，scan_table 验证表格渲染正常"],
    # I024
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I024", "横向滚动条拖拽定位测试", "低级", "拖拽横向滚动条可精确定位到任意列", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 拖拽横向滚动条到大约中间位置\n2. 释放鼠标",
     "中间位置",
     "表格滚动到对应位置，中间位置的列表头完整显示在视口中央，无截断，滚动距离与拖拽距离成比例",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 计算滚动条中间位置坐标，click_xy 拖拽滚动条到中间，scan_table 验证视口中央列表头完整可见"],
    # I025
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I025", "可跳转列单元格点击跳转验证", "低级", "缴库单号等可点击列点击后正确跳转详情页", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 点击第一行缴库单号列的单元格链接",
     "第1行缴库单号",
     "页面跳转到对应缴库单详情页，URL包含正确的单据ID，详情页显示的单据信息与列表行一致",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 tab/url 变化，get_table_values 获取第1行缴库单号值，click_table_cell 点击该单元格，observe_wait 验证URL包含单据ID，验证详情页数据匹配"],
    # I026
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I026", "可跳转列新窗口打开测试", "低级", "Ctrl+点击可跳转列在新标签页打开", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 按住Ctrl键点击第一行缴库单号列的单元格",
     "",
     "在新浏览器标签页打开缴库单详情页，原列表页保持不变未跳转，新标签页URL正确包含单据ID",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 触发 Ctrl+click 事件在缴库单号单元格，observe_start 监听 tab 变化，验证新标签页打开，验证原页面URL未变"],
    # I027
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I027", "表头筛选图标点击弹出筛选菜单", "低级", "点击列头筛选图标弹出对应列的筛选菜单", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 点击序号列头的筛选图标",
     "",
     "弹出该列的筛选下拉菜单，显示「升序」「降序」「按值筛选」等选项，菜单定位准确在筛选图标下方，无偏移",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "scan_table 获取筛选图标坐标，click_xy 点击筛选图标，run_js 检查 .vtable-filter-menu 元素存在且可见，验证菜单项完整"],
    # I028
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I028", "列表头点击升序排序功能", "低级", "点击列表头按该列升序排列，数据排序正确", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载，重量列数据无序",
     "1. 点击重量列表头进行升序排序",
     "",
     "表格按重量列从小到大排列，第一行重量值最小，最后一行最大，所有数据严格递增，相等值顺序可浮动",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击重量列表头，get_table_values 获取重量列数据，验证数组为升序排列：data[i] <= data[i+1] 对所有i成立"],
    # I029
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I029", "列表头点击降序排序功能", "低级", "再次点击列表头按该列降序排列，排序切换正确", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 重量列已按升序排序完成",
     "1. 再次点击重量列表头进行降序排序",
     "",
     "表格按重量列从大到小排列，第一行重量值最大，最后一行最小，排序箭头图标从向上变为向下",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 再次点击重量列表头，get_table_values 获取重量列数据，验证数组为降序排列：data[i] >= data[i+1] 对所有i成立"],
    # I030
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I030", "第三次点击表头取消排序功能", "低级", "第三次点击列表头取消排序恢复默认顺序", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 重量列已按降序排序完成",
     "1. 第三次点击重量列表头",
     "",
     "表格恢复为默认排序（通常按ID或制单时间降序），列头排序箭头消失，排序状态清空",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 第三次点击重量列表头，run_js 检查排序箭头元素不存在，get_table_values 验证重量列不再是严格升/降序（如有足够数据）"],
    # I031
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I031", "字符串列中文拼音排序正确性", "中级", "中文字符串列按拼音首字母排序，顺序正确", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 客户名称列包含「安踏」「宝马」「长虹」等中文数据",
     "1. 点击客户名称列表头进行升序排序",
     "",
     "客户名称按拼音首字母A→B→C顺序排列：「安踏」(A)在前，「宝马」(B)居中，「长虹」(C)在后，中文拼音排序逻辑正确",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击客户名称列表头，get_table_values 获取客户名称列，验证拼音顺序正确：安踏 < 宝马 < 长虹"],
    # I032
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I032", "日期列排序正确性验证", "中级", "日期时间列排序正确，跨年跨月数据顺序无误", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "排序",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 制单时间列包含2025年和2026年的数据",
     "1. 点击制单时间列表头进行降序排序",
     "",
     "制单时间按从新到旧排列，2026年的记录全部排在2025年之前，同一年内按月日排序，日期排序逻辑无误",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击制单时间列表头，get_table_values 获取制单时间列，验证所有2026年记录排在2025年记录之前"],
    # I033
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I033", "排序后分页数据一致性", "中级", "排序后翻页，排序规则保持一致，无重复无遗漏", MODULE_LEVEL1, MODULE_LEVEL2, "数据一致性测试", "排序+分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 重量列已按降序排序，结果多于1页",
     "1. 记录第1页最后一行的重量值\n2. 点击「下一页」按钮",
     "",
     "第2页第一行的重量值 <= 第1页最后一行的重量值，跨页排序保持连续，无数据重复，无数据遗漏，全局排序正确",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击重量列降序，get_table_values 获取第1页最后一行重量值W1，click_xy 点击下一页，获取第2页第一行重量值W2，验证W2 <= W1"],
    # I034
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I034", "表格行悬停高亮效果测试", "低级", "鼠标悬停在表格行上时，该行高亮显示", MODULE_LEVEL1, MODULE_LEVEL2, "UI测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 鼠标移动到第5行记录上悬停",
     "",
     "第5行背景色变为高亮色（通常为浅灰或浅蓝色），鼠标移开后恢复原色，高亮效果流畅无卡顿",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 在第5行触发 mouseover 事件，run_js 检查该行 background-color 发生变化，触发 mouseout 后验证颜色恢复"],
    # I035
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I035", "表格行双击事件测试", "低级", "双击表格行可快捷打开详情（如设计有此功能）", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载，支持双击行打开详情",
     "1. 鼠标双击第3行记录的任意位置",
     "",
     "双击后打开该记录的详情弹窗/页面，效果与单击缴库单号链接一致，无延迟无重复触发",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 modal/tab 变化，run_js 在第3行触发 dblclick 事件，observe_wait 验证详情页打开"],
    # I036
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I036", "大数量表格虚拟滚动性能测试", "中级", "超过1000行数据时虚拟滚动性能流畅", MODULE_LEVEL1, MODULE_LEVEL2, "性能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 表格加载1000+条记录，启用虚拟滚动",
     "1. 快速滚动表格从第1行到第1000行\n2. 观察滚动流畅度和渲染速度",
     "1000+行",
     "滚动过程流畅无明显卡顿，FPS保持在30以上，滚动时白屏时间 < 100ms，滚动停止后数据立即渲染完成，内存占用稳定",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 快速触发垂直滚动从0到最大，计时开始，计时结束验证总耗时 < 3秒，scan_table 验证目标行数据渲染完成，run_js 检查控制台无性能警告"],
    # I037
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I037", "表格单元格内容复制功能", "低级", "表格文本内容可选中复制", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 数据表格已加载",
     "1. 鼠标选中第2行客户名称单元格的文本\n2. 按Ctrl+C复制\n3. 粘贴到记事本",
     "",
     "文本可正常选中高亮，复制成功，粘贴内容与原单元格文本完全一致，无多余空格或格式丢失",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 模拟文本选择和复制操作，读取剪贴板内容，验证与 get_table_values 获取的客户名称值完全一致"],
    # I038
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I038", "表格空状态显示测试", "低级", "无数据时表格显示友好空状态", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 筛选条件设置为无结果（如未来日期）",
     "1. 设置筛选条件使查询结果为空\n2. 点击查询按钮",
     "无匹配数据",
     "表格区域显示友好空状态插画 + 「暂无数据」提示文字，表头仍可见（或按设计），布局居中对齐，无报错信息",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "设置无结果筛选条件，click_xy 点击查询，scan_table 验证空状态，dom_tree 检查「暂无数据」文本存在，验证布局居中无错乱"],
    # I039
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I039", "表格列显示自定义配置测试", "中级", "表格列设置功能可自定义显示/隐藏列", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 支持列设置功能",
     "1. 点击列设置按钮\n2. 取消勾选「备注」列\n3. 确认设置",
     "隐藏备注列",
     "表格中备注列消失不见，列设置中备注选项为未勾选，其他列保持显示，刷新页面后自定义列配置保持（如有记忆功能）",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击列设置按钮，click 取消勾选备注列，click 确认，scan_table 验证备注列表头不存在，刷新后再次验证"],
    # I040
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_I040", "表格列拖拽调整顺序测试", "中级", "拖拽列表头可调整列顺序，数据对应正确", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "表格交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 支持列拖拽排序",
     "1. 拖拽「销售订单」列表头到「客户名称」列之后\n2. 释放鼠标",
     "列顺序调整",
     "列顺序变为「缴入仓库 → 客户名称 → 销售订单」，数据与列头正确对应，销售订单列数据未与其他列混淆，刷新后顺序保持",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 触发列拖拽事件，将销售订单列拖到客户名称列之后，scan_table 验证列顺序调整正确，验证销售订单列数据与调整前一致"],
])

# ========== P001-P020: 页面级与端到端测试 ==========
test_cases.extend([
    # P001
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P001", "页面首次加载数据表格渲染验证", "高级", "进入生产缴库明细页面后表格正常加载数据", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "页面加载",
     "1. 已登录 SCM 系统\n2. 在系统首页或其他模块页面",
     "1. 从左侧菜单点击「生产缴库明细」进入页面\n2. 观察页面加载过程",
     "",
     "页面加载完成后数据表格显示缴库记录，表头列完整（26列有效数据列），无报错弹窗，无空白表格，首屏加载时间 < 3秒",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "enter_module 进入模块，计时开始，observe_wait 等待加载，scan_table 验证列数量 >= 26，验证总加载时间 < 3秒"],
    # P002
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P002", "筛选区内联模式显示验证", "中级", "筛选区以内联模式显示，所有筛选字段在页面内可见", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "页面布局",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面",
     "1. 观察筛选区显示模式\n2. 如为弹窗模式则切换为内联模式",
     "",
     "筛选区所有25个字段在页面内直接显示，点击筛选按钮不弹出高级搜索弹窗，显示「收起▲」按钮，布局整齐无重叠",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "dom_tree 检查 .page-query 结构，检查是否有 .legions-pro-quick-filter-remaining 元素存在，验证按钮文本为「收起▲」，scan_filter_fields 验证字段数量=25"],
    # P003
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P003", "筛选区收起/展开交互测试", "低级", "点击收起/展开按钮可切换筛选区显示状态", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "页面布局",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 筛选区当前为展开状态显示所有字段",
     "1. 点击「收起▲」按钮\n2. 观察筛选区变化\n3. 再次点击「展开▼」按钮",
     "",
     "点击收起后筛选区仅显示前N个字段，按钮文本变为「展开▼」；再次点击后恢复显示所有字段，按钮文本变回「收起▲」",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click_xy 点击收起按钮，dom_tree 验证按钮文本变为「展开▼」，再次 click_xy 点击，验证按钮文本变回「收起▲」，验证所有字段恢复显示"],
    # P004
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P004", "分页控件功能完整验证", "中级", "分页所有按钮功能正常，页码计算正确", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 共有3页数据，当前在第1页",
     "1. 点击「下一页」→ 验证到第2页\n2. 点击「上一页」→ 验证回到第1页\n3. 点击页码「3」→ 验证到第3页\n4. 点击「首页」→ 验证到第1页\n5. 点击「末页」→ 验证到第3页",
     "3页数据",
     "所有分页按钮功能正常，页码跳转正确，每页记录数 = 页大小，第3页记录数 <= 页大小，总页数和总条数显示正确",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "依次点击各分页按钮，每次 scan_table 记录当前页号和记录数，验证所有跳转正确，验证总条数文字显示正确"],
    # P005
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P005", "每页显示条数切换功能", "低级", "切换每页显示条数，表格和分页更新正确", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 当前每页显示20条",
     "1. 点击每页显示条数下拉\n2. 选择「50条/页」",
     "50条/页",
     "表格显示50条记录，分页控件总页数自动重新计算，当前页保持在第1页，数据无重复无遗漏",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 打开每页条数下拉，click 选择50条，scan_table 验证当前页记录数=50，dom_tree 验证总页数已更新减少"],
    # P006
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P006", "跳转到指定页码功能", "低级", "输入指定页码回车可直接跳转", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 共有5页数据，当前在第1页",
     "1. 点击页码输入框\n2. 输入「3」\n3. 按回车键",
     "跳转到第3页",
     "页面立即跳转到第3页，输入框值更新为3，表格显示第3页数据，分页控件高亮第3页按钮",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "click 点击页码输入框，input 输入3并按回车，scan_table 验证当前页为第3页，dom_tree 验证第3页按钮高亮状态"],
    # P007
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P007", "页码输入边界值测试", "低级", "输入超出范围的页码，系统正确处理", MODULE_LEVEL1, MODULE_LEVEL2, "边界值测试", "分页",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 共有5页数据",
     "1. 在页码输入框输入「10」\n2. 按回车键\n3. 在页码输入框输入「0」\n4. 按回车键",
     "页码=10 和 0",
     "输入10时自动调整到最大页码5并跳转第5页；输入0时自动调整到1并跳转第1页；无报错，友好提示或静默修正",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "input 输入10按回车，scan_table 验证当前页=5，input 输入0按回车，scan_table 验证当前页=1，observe_wait 验证无报错"],
    # P008
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P008", "页面刷新状态保持测试", "中级", "刷新页面后，筛选条件、页码、列宽等状态保持", MODULE_LEVEL1, MODULE_LEVEL2, "状态一致性测试", "状态保持",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 已设置缴入仓库=样品仓，在第2页，缴入仓库列宽已调至400px",
     "1. 按F5刷新浏览器页面",
     "",
     "刷新后自动恢复：缴入仓库=样品仓筛选条件保持，当前页码仍为第2页，缴入仓库列宽仍为400px，用户无需重新设置",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "设置好所有状态，run_js 刷新页面，scan_filter_fields 验证筛选条件保持，scan_table 验证页码和列宽保持"],
    # P009
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P009", "面包屑导航跳转测试", "低级", "点击面包屑可正确跳转到上级页面", MODULE_LEVEL1, MODULE_LEVEL2, "功能测试", "导航",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 面包屑显示：首页 > 生产管理 > 生产缴库明细",
     "1. 点击面包屑中的「生产管理」",
     "",
     "正确跳转到生产管理模块首页，URL和页面内容与直接从菜单点击进入一致",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "observe_start 监听 url 变化，click 点击面包屑「生产管理」文本，observe_wait 验证URL变为生产管理首页地址"],
    # P010
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P010", "标签页多开隔离测试", "中级", "同时打开两个生产缴库明细标签页，状态独立不干扰", MODULE_LEVEL1, MODULE_LEVEL2, "隔离性测试", "多标签",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面（标签页A）",
     "1. Ctrl+点击菜单在新标签页打开生产缴库明细（标签页B）\n2. 在标签页B设置缴入仓库=样品仓筛选\n3. 切换回标签页A",
     "双标签页",
     "标签页A的筛选条件仍为空，未被标签页B的操作影响，两个标签页状态完全独立，互不干扰",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js Ctrl+click 打开新标签页B，在标签页B设置筛选条件，切换回标签页A，scan_filter_fields 验证标签页A筛选条件为空"],
    # P011
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P011", "浏览器前进后退状态恢复", "中级", "使用浏览器前进后退按钮，页面状态正确恢复", MODULE_LEVEL1, MODULE_LEVEL2, "状态一致性测试", "导航",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面，状态S1（无筛选，第1页）",
     "1. 设置缴入仓库=成品仓，到第2页（状态S2）\n2. 点击浏览器后退按钮\n3. 验证回到状态S1\n4. 点击浏览器前进按钮\n5. 验证回到状态S2",
     "",
     "后退后正确恢复S1状态，前进后正确恢复S2状态，URL参数与页面状态一致，无历史栈错乱",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "设置S2状态，run_js 触发 history.back()，验证S1状态，run_js 触发 history.forward()，验证S2状态"],
    # P012
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P012", "页面权限控制测试", "高级", "无权限用户访问页面显示正确提示", MODULE_LEVEL1, MODULE_LEVEL2, "权限测试", "安全",
     "1. 以无生产缴库明细权限的账号登录 SCM 系统",
     "1. 直接在地址栏输入生产缴库明细页面URL访问",
     "",
     "系统正确识别无权限，显示「您暂无权限访问该页面」提示，提供返回首页按钮，不泄露任何业务数据",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "（需要切换测试账号）run_js 直接导航到目标URL，dom_tree 验证权限提示文本存在，scan_table 验证无业务数据表格渲染"],
    # P013
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P013", "并发编辑数据冲突检测", "高级", "两个用户同时编辑同一条记录，系统提示冲突", MODULE_LEVEL1, MODULE_LEVEL2, "并发测试", "数据一致性",
     "1. 用户A和用户B同时打开同一条生产缴库单详情页\n2. 两人均看到相同的原始数据",
     "1. 用户A修改备注为「AAA」并保存\n2. 用户B稍后修改备注为「BBB」并保存",
     "并发编辑",
     "用户B保存时系统弹出「该记录已被其他用户修改，请刷新后重试」提示，B的修改被阻止，数据保持A修改后的值，无数据覆盖丢失",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "（需要两个浏览器会话）用户A执行修改保存，用户B在A之后保存，observe_wait 验证用户B看到冲突提示，验证最终数据为A修改的值"],
    # P014
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P014", "断网离线降级处理", "中级", "网络断开时操作显示友好提示，数据不丢失", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "离线处理",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面\n3. 手动断开网络连接",
     "1. 点击查询按钮\n2. 尝试其他操作",
     "模拟断网",
     "系统弹出「网络连接已断开，请检查网络后重试」友好提示，操作被中止，前端状态不崩溃，网络恢复后可正常继续操作",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 模拟 offline 事件，click_xy 点击查询按钮，observe_wait 验证离线提示弹出，验证无JS报错，run_js 恢复 online 后验证操作恢复正常"],
    # P015
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P015", "响应式布局适配测试", "中级", "不同屏幕尺寸下页面布局自适应调整", MODULE_LEVEL1, MODULE_LEVEL2, "兼容性测试", "响应式",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面",
     "1. 调整浏览器窗口宽度为1920px（桌面）\n2. 调整为1366px（笔记本）\n3. 调整为768px（平板）",
     "三档宽度",
     "1920px下所有列完整显示；1366px下表头出现横向滚动条；768px下筛选区自动折叠，表格横向滚动可用，按钮自适应宽度，无元素溢出",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 分别设置窗口宽度为1920、1366、768，每次 scan_page_elements 验证布局正常，无元素溢出视口，无重叠，关键按钮均可见"],
    # P016
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P016", "键盘快捷键操作支持", "低级", "常用操作支持键盘快捷键", MODULE_LEVEL1, MODULE_LEVEL2, "可访问性测试", "交互",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面",
     "1. 按Tab键遍历所有可交互元素\n2. 在筛选输入框按Enter触发查询\n3. 按Esc关闭弹窗",
     "",
     "Tab键遍历顺序合理（筛选区→按钮→表格→分页），无跳漏；输入框按Enter正确触发查询；Esc键可关闭当前弹窗；符合无障碍规范",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 模拟Tab键遍历，记录可聚焦元素数量和顺序，run_js 在输入框触发Enter键，验证查询执行，run_js 触发Esc键，验证弹窗关闭"],
    # P017
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P017", "长时间会话过期处理", "高级", "会话过期后操作正确跳转到登录页", MODULE_LEVEL1, MODULE_LEVEL2, "异常测试", "会话",
     "1. 已登录 SCM 系统\n2. 在生产缴库明细页面停留超过会话超时时间",
     "1. 点击查询按钮",
     "会话已过期",
     "系统检测到会话过期，弹出「登录已过期，请重新登录」提示，3秒后自动跳转到登录页面，当前页面操作被安全阻止",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "run_js 清除本地会话token模拟过期，click_xy 点击查询，observe_wait 验证过期提示弹出，验证3秒后跳转到登录页"],
    # P018
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P018", "端到端完整业务流程测试", "高级", "从筛选→查看详情→打印的完整业务流", MODULE_LEVEL1, MODULE_LEVEL2, "端到端测试", "完整流程",
     "1. 已登录 SCM 系统\n2. 在系统首页",
     "1. 进入生产缴库明细模块\n2. 筛选缴入仓库=成品仓，审批状态=已通过\n3. 点击第一条记录的缴库单号查看详情\n4. 核实详情页数据与列表一致\n5. 在详情页点击打印按钮\n6. 确认打印预览关闭后返回列表",
     "完整流程",
     "整个流程无报错，数据在各环节保持一致，打印预览正确，返回列表后筛选条件仍保持，用户体验流畅，无中断",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "按步骤依次执行，每步验证正确，记录第一条数据的关键字段值，详情页验证值一致，返回列表后验证筛选条件保持"],
    # P019
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P019", "数据审计日志记录验证", "高级", "关键操作正确记录审计日志", MODULE_LEVEL1, MODULE_LEVEL2, "安全测试", "审计",
     "1. 已登录 SCM 系统管理员账号\n2. 在生产缴库明细页面",
     "1. 执行一次导出操作\n2. 进入审计日志模块查询",
     "导出操作",
     "审计日志中正确记录本次导出操作：操作人、操作时间、操作类型=导出、模块=生产缴库明细、IP地址等信息完整",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "记录操作时间和操作人账号，click_xy 点击导出，进入审计日志模块，查询该时间段该账号的操作日志，验证导出操作记录存在且字段完整"],
    # P020
    [f"{ENTERPRISE_PREFIX}_{MODULE_PINYIN}_P020", "连续8小时稳定性测试", "高级", "页面长时间打开后操作仍正常，无内存泄漏", MODULE_LEVEL1, MODULE_LEVEL2, "稳定性测试", "长稳",
     "1. 已登录 SCM 系统\n2. 生产缴库明细页面保持打开8小时",
     "1. 8小时后执行：查询→排序→翻页→导出",
     "8小时长稳",
     "所有操作响应正常，响应时间与页面刚打开时无显著差异（< 2倍），无内存溢出报错，无界面卡顿，控制台无持续性错误日志",
     "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), "",
     "（自动化长稳测试）记录初始内存占用和响应时间，8小时后重复相同操作，比较响应时间增长 < 2倍，比较内存增长 < 100MB，验证无持续性错误"],
])

# 写入用例数据
for row_data in test_cases:
    ws1.append(row_data)

# 应用单元格格式
for row_idx in range(2, ws1.max_row + 1):
    for col_idx in range(1, ws1.max_column + 1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx in (1, 2, 3, 5, 6, 7, 8, 13, 14, 15, 16, 17):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    apply_priority_style(ws1.cell(row=row_idx, column=3))

# 列宽设置
COL_WIDTHS = [18, 42, 12, 42, 18, 18, 18, 12, 50, 44, 44, 50, 10, 10, 12, 12, 12, 0, 58]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[chr(64 + i)].width = w

ws1.freeze_panes = "E2"
ws1.auto_filter.ref = f"A1:{ws1.cell(1, ws1.max_column).column_letter}{ws1.max_row}"

# ===================== Sheet 2: 测试数据与统计 =====================
ws2 = wb.create_sheet(title="测试数据")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11, name="微软雅黑")
header2_font = Font(bold=True, size=10, name="微软雅黑")

def write_section(ws, start_row, title, headers, data_rows):
    col_count = len(headers)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = section_fill
    title_cell.font = section_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    hdr_row = start_row + 1
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = header2_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for row_idx, row_data in enumerate(data_rows, hdr_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return hdr_row + 1 + len(data_rows) + 2

row = 1
# 测试用例统计
row = write_section(ws2, row, "测试用例统计汇总",
    ["用例类型", "数量", "覆盖范围"],
    [
        ["筛选查询类 (F001-F025)", "25", "单字段筛选、边界值、异常、多条件组合、一致性、幂等性"],
        ["按钮交互类 (I001-I015)", "15", "打印/导出/批量打印/添加/流程设置的正向/异常/性能场景"],
        ["表格操作类 (I016-I040)", "25", "复选框、列宽、滚动、排序、跳转、虚拟滚动、列设置、空状态"],
        ["页面级与端到端 (P001-P020)", "20", "页面加载、分页、导航、状态保持、权限、并发、稳定性"],
        ["总计", "85", "覆盖功能测试、边界测试、异常测试、性能测试、安全测试、兼容性测试"],
    ])

# 测试类型矩阵
row = write_section(ws2, row, "测试类型覆盖矩阵",
    ["测试类型", "说明", "用例数量", "优先级"],
    [
        ["功能测试", "验证核心业务功能正确性", "45", "最高"],
        ["边界值测试", "验证输入/操作边界的系统处理", "12", "高"],
        ["异常测试", "验证异常场景下的系统健壮性", "10", "高"],
        ["组合条件测试", "验证多条件组合逻辑正确性", "3", "中"],
        ["数据一致性测试", "验证跨操作/跨页面数据一致", "8", "高"],
        ["幂等性测试", "验证重复操作结果一致", "3", "中"],
        ["性能测试", "验证系统响应速度和资源占用", "5", "中"],
        ["稳定性测试", "验证长时间运行可靠性", "1", "中"],
        ["并发测试", "验证多用户/多操作并发正确性", "1", "高"],
        ["权限测试", "验证权限控制和数据安全", "2", "最高"],
        ["兼容性测试", "验证不同环境下表现一致", "1", "中"],
        ["可访问性测试", "验证键盘操作等无障碍支持", "1", "低"],
    ])

row = write_section(ws2, row, "3.1 筛选用例测试数据配置",
    ["字段名称", "操作符", "测试值1", "测试值2", "测试值3", "预期结果"],
    [
        ["缴入仓库", "等于", "样品仓", "成品仓", "半成品仓", "仅显示对应仓库数据"],
        ["制单时间", "范围", "2026-07-01 ~ 2026-07-15", "2026-07-01 ~ 2026-07-01", "2026-07-10 ~ 2026-07-01(非法)", "日期范围内数据/同一天/提示开始>结束"],
        ["审批状态", "等于", "已通过", "审批中", "已驳回", "仅显示对应状态数据"],
        ["重量", "小于/大于", "< 100", "> 50", "< 0(负数)", "数值比较/负数处理正确"],
        ["商品名称", "包含", "轴承", "特殊字符'\"<", "超长200字符", "字符串匹配/防注入/截断正确"],
    ])

row = write_section(ws2, row, "3.2 筛选字段完整列表",
    ["序号", "字段名称", "输入方式", "操作符支持", "备注"],
    [
        ["1", "重量", "数值输入", "等于/不等于/大于/小于/大于等于/小于等于", ""],
        ["2", "应生产数量", "数值输入", "等于/不等于/大于/小于/大于等于/小于等于", ""],
        ["3", "批次号", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["4", "缴入仓库", "下拉多选", "等于/不等于", "16个仓库选项 + 全选"],
        ["5", "销售订单", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["6", "客户名称", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["7", "客户编码", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["8", "缴库单号", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["9", "缴库订单号", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["10", "审批状态", "下拉单选", "等于", "审批中/已通过/已驳回/已删除"],
        ["11", "审批人", "下拉选择", "等于", "用户列表"],
        ["12", "商品识别码", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["13", "商品名称", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["14", "商品规格", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["15", "加工方式", "下拉单选", "等于", "自制/外购/外发加工/委外加工"],
        ["16", "中类名称", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["17", "关联单号", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["18", "生产部门", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["19", "备注", "文本输入", "等于/包含/开头是/结尾是", ""],
        ["20", "制单时间", "日期范围", "介于/早于/晚于", "日期选择器"],
        ["21", "缴库时间", "日期范围", "介于/早于/晚于", "日期选择器"],
        ["22", "缴库数量", "数值输入", "等于/不等于/大于/小于/大于等于/小于等于", ""],
        ["23", "经办人", "下拉选择", "等于", "用户列表"],
        ["24", "入库状态", "下拉单选", "等于", "已缴库/待缴库"],
        ["25", "创建人", "下拉选择", "等于", "用户列表"],
    ])

row = write_section(ws2, row, "3.3 工具栏按钮一览表",
    ["按钮名称", "快捷键", "是否需勾选行", "权限要求", "功能说明"],
    [
        ["导 出", "无", "否", "生产缴库明细-导出", "导出当前筛选结果为Excel"],
        ["打 印", "无", "是", "生产缴库明细-打印", "打印选中记录"],
        ["批量打印", "无", "是", "生产缴库明细-批量打印", "批量打印选中的多条记录"],
        ["添加生产缴库单", "无", "是", "生产缴库明细-新增", "基于选中记录新增缴库单"],
        ["流程设置", "无", "是", "生产缴库明细-流程配置", "配置选中记录的审批流程"],
        ["查 询", "Enter", "否", "生产缴库明细-查看", "按筛选条件查询数据"],
        ["重 置", "Esc", "否", "生产缴库明细-查看", "清空所有筛选条件"],
        ["列设置", "无", "否", "生产缴库明细-查看", "自定义表格显示列"],
    ])

row = write_section(ws2, row, "3.4 VTable表格列定义一览表",
    ["列序号", "列标题", "数据类型", "可排序", "可筛选", "可跳转", "固定列", "说明"],
    [
        ["1", "复选框", "boolean", "否", "否", "否", "是", "首列全选/行勾选"],
        ["2", "序号", "integer", "是", "是", "否", "否", "行号显示"],
        ["3", "缴库单号", "string", "是", "是", "是", "否", "点击跳转到缴库单详情"],
        ["4", "缴库订单号", "string", "是", "是", "是", "否", "点击跳转到订单详情"],
        ["5", "审批状态", "enum", "是", "是", "是", "否", "点击查看审批流程图"],
        ["6", "审批人", "string", "是", "是", "否", "否", "最后审批人姓名"],
        ["7", "缴入仓库", "enum", "是", "是", "否", "否", "16个仓库枚举"],
        ["8", "销售订单", "string", "是", "是", "是", "否", "关联销售订单号"],
        ["9", "客户名称", "string", "是", "是", "否", "否", "客户全称"],
        ["10", "客户编码", "string", "是", "是", "否", "否", "客户系统编码"],
        ["11", "应生产数量", "integer", "是", "是", "否", "否", "生产订单数量"],
        ["12", "批次号", "string", "是", "是", "否", "否", "生产批次"],
        ["13", "重量", "decimal", "是", "是", "否", "否", "单位:千克"],
        ["14", "商品识别码", "string", "是", "是", "是", "否", "点击跳转到商品档案"],
        ["15", "商品规格", "string", "是", "是", "否", "否", "商品规格型号"],
        ["16", "商品名称", "string", "是", "是", "否", "否", "商品全称"],
        ["17", "加工方式", "enum", "是", "是", "是", "否", "自制/外购/外发加工/委外加工"],
        ["18", "中类名称", "string", "是", "是", "否", "否", "商品分类中类"],
        ["19", "关联单号", "string", "是", "是", "是", "否", "关联的上游单据号"],
        ["20", "生产部门", "string", "是", "是", "否", "否", "负责生产的部门"],
        ["21", "备注", "string", "是", "是", "否", "否", "备注说明"],
        ["22", "制单时间", "datetime", "是", "是", "是", "否", "单据创建时间"],
        ["23", "缴库时间", "datetime", "是", "是", "是", "否", "实际入库时间"],
        ["24", "缴库数量", "integer", "是", "是", "否", "否", "实际入库数量"],
        ["25", "经办人", "string", "是", "是", "否", "否", "缴库操作人"],
        ["26", "入库状态", "enum", "是", "是", "是", "否", "已缴库/待缴库"],
    ])

# ===================== 保存文件 =====================
filename = f"测试用例_生产管理_生产缴库明细_完整版_{date.today().isoformat()}.xlsx"
filepath = os.path.join(OUTPUT_DIR, filename)
wb.save(filepath)

print(f"✅ 测试用例生成完成！")
print(f"📁 文件路径: {os.path.abspath(filepath)}")
print(f"📊 用例数量: {len(test_cases)} 条")
print(f"📋 Sheet1: 测试用例（{len(HEADERS_19)}列完整格式）")
print(f"📋 Sheet2: 测试数据配置 + 统计汇总 + 字段说明 + 按钮列表 + 表格列定义")