#!/usr/bin/env python3
# Excel 测试用例导出模板
# 用法：在 eval kernel 中执行，test_cases 变量由 AI 在 Phase 4 填入

# ============================================================
# 可配置变量 — Phase 4 执行时按实际项目修改
# ============================================================
ENTERPRISE_PREFIX = "NB"
DEFAULT_AUTHOR    = "Hooplus1ce"
MODULE_NAME       = "生产管理_制造排产"
OUTPUT_DIR        = "."

# ============================================================
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ===================== Sheet 1: 测试用例 =====================
ws1 = wb.active
ws1.title = "测试用例"

HEADERS_18 = [
    "用例编号", "用例标题", "级别", "验证点",
    "一级模块", "二级模块", "测试类型", "功能",
    "前置条件", "测试步骤", "测试数据", "预期结果",
    "测试结果", "执行人", "执行时间", "编写人", "编写时间", "备注"
]
ws1.append(HEADERS_18)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

PRIORITY_STYLES = {
    "高级": (PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid"),
           Font(bold=True, color="9C0006", size=10)),
    "中级": (PatternFill(start_color="FFE8D0", end_color="FFE8D0", fill_type="solid"),
           Font(bold=True, color="9C6500", size=10)),
    "低级": (PatternFill(start_color="FFF5C0", end_color="FFF5C0", fill_type="solid"),
           Font(bold=True, color="806000", size=10)),
}

def apply_priority_style(cell):
    style = PRIORITY_STYLES.get(cell.value)
    if style:
        cell.fill, cell.font = style

# ============================================================
# test_cases — AI 在 Phase 4 填入实际用例数据
# 格式：每行 18 个字段，对应 HEADERS_18 顺序
# 多行内容用 \n 连接，导出后自动换行显示
# ============================================================
test_cases = [
    # 示例：
    # [f"{ENTERPRISE_PREFIX}_MOD_001", "用例标题", "中级", "验证点",
    #  "一级模块", "二级模块", "功能测试", "查询",
    #  "前置条件", "测试步骤", "测试数据", "预期结果",
    #  "", "", "", DEFAULT_AUTHOR, date.today().isoformat(), ""],
]

for row_data in test_cases:
    ws1.append(row_data)
    row_idx = ws1.max_row
    for cell_idx, cell in enumerate(ws1[row_idx], 1):
        cell.border = thin_border
        if cell_idx in (2, 4, 9, 10, 11, 12):  # B/D/I/J/K/L 列左对齐
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_priority_style(ws1.cell(row=row_idx, column=3))

COL_WIDTHS = [18, 42, 12, 42, 18, 18, 18, 12, 50, 44, 44, 50, 10, 10, 12, 12, 12, 0]
for i, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[chr(64 + i)].width = w

ws1.freeze_panes = "E2"
ws1.auto_filter.ref = f"A1:R{ws1.max_row}"

# ===================== Sheet 2: 测试数据 =====================
ws2 = wb.create_sheet(title="测试数据")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11, name="微软雅黑")
header2_font = Font(bold=True, size=10, name="微软雅黑")

def write_section(ws, start_row, title, headers, data_rows):
    col_count = len(headers)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = section_font
    title_cell.fill = section_fill
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    for c in range(1, col_count + 1):
        ws.cell(row=start_row, column=c).fill = section_fill
        ws.cell(row=start_row, column=c).border = thin_border
    hdr_row = start_row + 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font = header2_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_offset, row_data in enumerate(data_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=hdr_row + r_offset, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    return hdr_row + 1 + len(data_rows) + 2

row = 1
row = write_section(ws2, row, "3.1 测试数据配置",
    ["序号", "系统编号", "一级模块", "二级模块", "功能",
     "用例标题", "级别", "测试类型", "编写人", "编写日期"], [])

row = write_section(ws2, row, "3.2 筛选字段说明",
    ["字段名称", "输入方式", "对应测试数据", "说明"], [])

row = write_section(ws2, row, "3.3 各页签工具栏按钮一览表",
    ["状态页签", "按钮名称", "是否需要勾选行"], [])

row = write_section(ws2, row, "3.4 VTable列定义一览表",
    ["列标题", "字段名", "数据格式说明", "列头交互能力"], [])

# ===================== 保存 =====================
filename = f"测试用例_{MODULE_NAME}_{date.today().isoformat()}.xlsx"
os.makedirs(OUTPUT_DIR, exist_ok=True)
filepath = os.path.join(OUTPUT_DIR, filename)
try:
    wb.save(filepath)
    print(f"✅ 已生成: {os.path.abspath(filepath)}")
except PermissionError:
    fallback = os.path.join(".", filename)
    wb.save(fallback)
    print(f"⚠️ 降级保存至: {os.path.abspath(fallback)}")
